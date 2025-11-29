#!/usr/bin/env python3
from __future__ import annotations

"""draft_dump: dump Yahoo Fantasy league draft results.

This script depends on a prior run of ``league_dump`` for the same league.
It reads the latest processed league JSON via:

  exports/<league_key>/_meta/latest.json  →  league_dump.processed

and then fetches draft results from Yahoo to build:

  exports/<league_key>/draft_dump/
    raw/
      draftresults.<ISO>.json          Raw Yahoo payload
    processed/
      draft.<ISO>.json                 Normalized draft results JSON
    excel/
      draft.<ISO>.xlsx                 Optional Excel workbook (--to-excel)
    manifest/
      manifest.<ISO>.json              Manifest for this run

The _meta/latest.json file is updated with a ``draft_dump`` block alongside
existing modules (league_dump, standings_dump, transactions_dump, ...).
"""


import argparse
import json
import sys
import hashlib
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional

from src.auth.oauth import get_session
from src.config.env import get_export_dir
from src.util_time import RunTimestamps, make_run_timestamps

BASE_URL = "https://fantasysports.yahooapis.com/fantasy/v2"


# ---------------- HTTP helpers ----------------


def _fetch(endpoint: str) -> dict:
    """Fetch a Fantasy Sports API endpoint as JSON, raising for HTTP errors."""
    sess = get_session()
    url = f"{BASE_URL}/{endpoint}"
    r = sess.get(url, params={"format": "json"}, headers={"Accept": "application/json"})
    r.raise_for_status()
    return r.json()


def _fetch_draftresults(league_key: str) -> dict:
    """Fetch /league/{league_key}/draftresults as JSON."""
    endpoint = f"league/{league_key}/draftresults"
    return _fetch(endpoint)


# ---------------- league_dump context loading ----------------


@dataclass
class LeagueContext:
    league_key: str
    league_info: Dict[str, Any]
    teams: List[Dict[str, Any]]
    scoring: Dict[str, Any]


def _load_latest_league_dump(league_key: str) -> LeagueContext:
    """Load the latest processed league_dump JSON for this league.

    Error and exit if no league_dump entry is present, since draft_dump
    is designed to build on top of league_dump outputs.
    """
    root = get_export_dir() / league_key
    meta_dir = root / "_meta"
    latest_path = meta_dir / "latest.json"

    if not latest_path.exists():
        print(f"[ERROR] No _meta/latest.json for league {league_key}.")
        print("        Run league_dump first, then re-run draft_dump.")
        raise SystemExit(1)

    with latest_path.open("r", encoding="utf-8") as f:
        latest = json.load(f)

    ld = latest.get("league_dump")
    if not ld or "processed" not in ld:
        print(f"[ERROR] _meta/latest.json for {league_key} has no 'league_dump.processed' entry.")
        print("        Run league_dump first, then re-run draft_dump.")
        raise SystemExit(1)

    rel_processed = ld["processed"]
    processed_path = root / rel_processed

    if not processed_path.exists():
        print(f"[ERROR] league_dump processed file not found: {processed_path}")
        print("        Re-run league_dump for this league, then re-run draft_dump.")
        raise SystemExit(1)

    with processed_path.open("r", encoding="utf-8") as f:
        payload = json.load(f)

    league_info = payload.get("league_info") or {}
    teams = payload.get("teams") or []
    scoring = payload.get("scoring") or {}

    return LeagueContext(
        league_key=league_key,
        league_info=league_info,
        teams=teams,
        scoring=scoring,
    )


# ---------------- paths + manifest helpers ----------------


@dataclass
class DraftPaths:
    league_key: str
    root: Path
    meta_dir: Path
    raw_dir: Path
    processed_dir: Path
    excel_dir: Path
    manifest_dir: Path


def _prepare_draft_dirs(league_key: str) -> DraftPaths:
    """Prepare league-scoped export directories under exports/<league_key>/draft_dump/…"""
    base = get_export_dir()
    league_root = base / league_key
    meta_dir = league_root / "_meta"
    dd_root = league_root / "draft_dump"
    raw_dir = dd_root / "raw"
    processed_dir = dd_root / "processed"
    excel_dir = dd_root / "excel"
    manifest_dir = dd_root / "manifest"

    for d in (meta_dir, raw_dir, processed_dir, excel_dir, manifest_dir):
        d.mkdir(parents=True, exist_ok=True)

    return DraftPaths(
        league_key=league_key,
        root=league_root,
        meta_dir=meta_dir,
        raw_dir=raw_dir,
        processed_dir=processed_dir,
        excel_dir=excel_dir,
        manifest_dir=manifest_dir,
    )


def _sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _build_manifest_dict(
    module_name: str,
    league_key: str,
    paths: DraftPaths,
    run_ts: RunTimestamps,
    cli_args: Dict[str, Any],
    produced_paths: List[Path],
) -> Dict[str, Any]:
    files: Dict[str, Dict[str, Any]] = {}

    for abs_path in produced_paths:
        # Normalize to league-root-relative POSIX paths
        rel = abs_path.relative_to(paths.root).as_posix()
        stat = abs_path.stat()
        files[rel] = {
            "size_bytes": stat.st_size,
            "sha256": _sha256_file(abs_path),
        }

    return {
        "module": module_name,
        "league_key": league_key,
        "_generated_unix": run_ts.unix,
        "_generated_iso_utc": run_ts.iso_utc,
        "_generated_iso_local": run_ts.iso_local,
        "files": files,
        "cli_args": cli_args,
    }


def _write_manifest(paths: DraftPaths, run_ts: RunTimestamps, manifest_data: Dict[str, Any]) -> Path:
    out_path = paths.manifest_dir / f"manifest.{run_ts.iso_stamp}.json"
    with out_path.open("w", encoding="utf-8") as f:
        json.dump(manifest_data, f, indent=2, sort_keys=True)
    return out_path


def _update_latest(
    paths: DraftPaths,
    run_ts: RunTimestamps,
    raw_rel: str,
    processed_rel: str,
    excel_rel: Optional[str],
) -> Path:
    """Update _meta/latest.json with a 'draft_dump' block, preserving other modules' keys."""
    latest_path = paths.meta_dir / "latest.json"

    if latest_path.exists():
        with latest_path.open("r", encoding="utf-8") as f:
            data = json.load(f)
    else:
        data = {"league_key": paths.league_key}

    dd = data.get("draft_dump", {})
    dd["raw"] = raw_rel
    dd["processed"] = processed_rel
    if excel_rel is not None:
        dd["excel"] = excel_rel

    data["draft_dump"] = dd
    data["_updated_unix"] = run_ts.unix
    data["_updated_iso_utc"] = run_ts.iso_utc

    with latest_path.open("w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, sort_keys=True)

    return latest_path


# ---------------- draft normalization + Excel ----------------


def _normalize_draftresults(
    raw_payload: Dict[str, Any],
    league_ctx: LeagueContext,
    run_ts: RunTimestamps,
) -> Dict[str, Any]:
    """Normalize Yahoo draftresults payload into a tidy JSON structure."""
    fc = raw_payload.get("fantasy_content") or {}
    league_node = fc.get("league")

    if not isinstance(league_node, list) or len(league_node) < 2:
        raise ValueError("Unexpected draftresults payload shape: missing league list")

    meta = league_node[0] or {}
    dr_container = league_node[1].get("draft_results") or {}
    if not isinstance(dr_container, dict):
        raise ValueError("Unexpected draftresults payload shape: league[1].draft_results missing")

    league_key = meta.get("league_key") or league_ctx.league_info.get("league_key") or league_ctx.league_key
    league_name = meta.get("name") or league_ctx.league_info.get("name") or league_ctx.league_info.get("league_name")
    season = meta.get("season") or league_ctx.league_info.get("season")
    draft_status = meta.get("draft_status") or league_ctx.league_info.get("draft_status")
    num_teams = meta.get("num_teams") or league_ctx.league_info.get("num_teams") or len(league_ctx.teams)

    try:
        num_teams_int = int(num_teams) if num_teams is not None else None
    except (TypeError, ValueError):
        num_teams_int = None

    # Map team_key → team_name from league_dump teams.
    team_name_by_key: Dict[str, Optional[str]] = {}
    for t in league_ctx.teams:
        tk = t.get("team_key")
        if tk:
            team_name_by_key[tk] = t.get("name")

    results: List[Dict[str, Any]] = []
    for key, val in dr_container.items():
        if key == "count":
            continue
        dr = val.get("draft_result") or {}
        pick = dr.get("pick")
        rnd = dr.get("round")
        team_key = dr.get("team_key")
        player_key = dr.get("player_key")

        try:
            pick_int = int(pick) if pick is not None else None
        except (TypeError, ValueError):
            pick_int = None
        try:
            round_int = int(rnd) if rnd is not None else None
        except (TypeError, ValueError):
            round_int = None

        results.append(
            {
                "pick": pick_int,
                "round": round_int,
                "team_key": team_key,
                "team_name": team_name_by_key.get(team_key),
                "player_key": player_key,
            }
        )

    # Sort by overall pick then round for stable output.
    results.sort(key=lambda r: (r.get("pick") or 0, r.get("round") or 0))

    out: Dict[str, Any] = {
        "league_key": league_key,
        "league_name": league_name,
        "season": season,
        "num_teams": num_teams_int,
        "draft_status": draft_status,
        "fetched_unix": run_ts.unix,
        "fetched_iso_utc": run_ts.iso_utc,
        "fetched_iso_local": run_ts.iso_local,
        "count": len(results),
        "draft_results": results,
    }
    return out


def _write_excel(draft_processed: Dict[str, Any], path: Path) -> None:
    """Write an Excel workbook with ByRound, ByTeam, and RunInfo sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(
            "ERROR: openpyxl is not installed but --to-excel was requested.\n"
            "Install it with 'pip install openpyxl' and try again.",
            file=sys.stderr,
        )
        sys.exit(1)

    path.parent.mkdir(parents=True, exist_ok=True)

    results: List[Dict[str, Any]] = list(draft_processed.get("draft_results") or [])

    wb = Workbook()
    ws_round = wb.active
    ws_round.title = "ByRound"

    headers = ["pick", "round", "team_key", "team_name", "player_key"]

    def write_sheet(ws, rows: List[Dict[str, Any]]) -> None:
        ws.append(headers)
        # Freeze header row and enable auto-filter
        ws.freeze_panes = "A2"
        last_col_letter = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col_letter}1"

        for row in rows:
            ws.append([row.get(h) for h in headers])

        # Auto-size columns a bit
        for idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].width = 18

    # ByRound sheet
    by_round = sorted(results, key=lambda r: (r.get("round") or 0, r.get("pick") or 0))
    write_sheet(ws_round, by_round)

    # ByTeam sheet
    ws_team = wb.create_sheet("ByTeam")
    by_team = sorted(
        results,
        key=lambda r: (
            r.get("team_name") or r.get("team_key") or "",
            r.get("round") or 0,
            r.get("pick") or 0,
        ),
    )
    write_sheet(ws_team, by_team)

    # RunInfo sheet
    ws_info = wb.create_sheet("RunInfo")
    ws_info.append(
        [
            "league_key",
            "league_name",
            "season",
            "num_teams",
            "draft_status",
            "fetched_unix",
            "fetched_iso_utc",
            "fetched_iso_local",
        ]
    )
    ws_info.append(
        [
            draft_processed.get("league_key"),
            draft_processed.get("league_name"),
            draft_processed.get("season"),
            draft_processed.get("num_teams"),
            draft_processed.get("draft_status"),
            draft_processed.get("fetched_unix"),
            draft_processed.get("fetched_iso_utc"),
            draft_processed.get("fetched_iso_local"),
        ]
    )

    wb.save(path)


# ---------------- CLI + main ----------------


def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Dump Yahoo Fantasy league draft results.")
    g = p.add_mutually_exclusive_group(required=True)
    g.add_argument("--league-key", help="Full league key, e.g. 465.l.22607")
    g.add_argument("--league-id", type=int, help="League ID (use with --game)")

    p.add_argument("--game", default="nhl", help="Game code/key (default: nhl)")
    p.add_argument("--pretty", action="store_true", help="Pretty-print JSON outputs")
    p.add_argument("--to-excel", action="store_true", help="Also write an Excel workbook")

    return p.parse_args()


def _resolve_league_key(args: argparse.Namespace) -> str:
    if args.league_key:
        return args.league_key
    return f"{args.game}.l.{args.league_id}"


def main() -> None:
    args = _parse_args()
    league_key = _resolve_league_key(args)

    # Load league context from league_dump (required for team names, season, etc.).
    league_ctx = _load_latest_league_dump(league_key)

    # Prepare directories and timestamps for this run.
    paths = _prepare_draft_dirs(league_key)
    run_ts = make_run_timestamps()
    iso_stamp = run_ts.iso_stamp

    # Fetch raw draftresults payload.
    raw_payload = _fetch_draftresults(league_key)

    # Write raw payload snapshot.
    raw_path = paths.raw_dir / f"draftresults.{iso_stamp}.json"
    with raw_path.open("w", encoding="utf-8") as f:
        json.dump(raw_payload, f, ensure_ascii=False, indent=2 if args.pretty else None)
    print(f"Wrote raw draftresults payload: {raw_path}")

    # Normalize into processed JSON.
    processed = _normalize_draftresults(raw_payload=raw_payload, league_ctx=league_ctx, run_ts=run_ts)
    processed_path = paths.processed_dir / f"draft.{iso_stamp}.json"
    with processed_path.open("w", encoding="utf-8") as f:
        json.dump(processed, f, ensure_ascii=False, indent=2 if args.pretty else None)
    print(f"Wrote processed draft JSON: {processed_path}")

    # Optional Excel workbook.
    excel_path: Optional[Path] = None
    if args.to_excel:
        excel_path = paths.excel_dir / f"draft.{iso_stamp}.xlsx"
        _write_excel(processed, excel_path)
        print(f"Wrote draft Excel workbook: {excel_path}")

    # Manifest + latest.json update.
    produced_paths: List[Path] = [raw_path, processed_path]
    if excel_path is not None:
        produced_paths.append(excel_path)

    cli_args: Dict[str, Any] = {
        "league_key": league_key,
        "league_id": getattr(args, "league_id", None),
        "game": getattr(args, "game", None),
        "pretty": bool(getattr(args, "pretty", False)),
        "to_excel": bool(getattr(args, "to_excel", False)),
    }

    manifest_data = _build_manifest_dict(
        module_name="draft_dump",
        league_key=league_key,
        paths=paths,
        run_ts=run_ts,
        cli_args=cli_args,
        produced_paths=produced_paths,
    )
    manifest_path = _write_manifest(paths, run_ts, manifest_data)
    print(f"Wrote manifest: {manifest_path}")

    raw_rel = raw_path.relative_to(paths.root).as_posix()
    processed_rel = processed_path.relative_to(paths.root).as_posix()
    excel_rel = excel_path.relative_to(paths.root).as_posix() if excel_path is not None else None
    latest_path = _update_latest(paths, run_ts, raw_rel=raw_rel, processed_rel=processed_rel, excel_rel=excel_rel)
    print(f"Updated latest.json: {latest_path}")


if __name__ == "__main__":  # pragma: no cover
    main()
