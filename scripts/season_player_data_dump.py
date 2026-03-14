#!/usr/bin/env python3
from __future__ import annotations

"""season_player_data_dump: fetch per-player season payloads and write one JSON file per player.

Writes to `exports/playerdata/<season>/<player_key>.json` and produces a stat-id map
at `exports/playerdata/<season>/stat_id_map.<iso>.json`.

This reuses the same Yahoo endpoint pattern used by `league_players_dump` but
outputs per-player JSON files suitable for later game aggregation.
"""

import argparse
import json
import math
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

from src.auth.oauth import get_session
from src.config.env import get_export_dir
from src.util_time import make_run_timestamps, RunTimestamps

from src.yahoo.api_error import handle_api_error

API_BASE = "https://fantasysports.yahooapis.com/fantasy/v2"
BATCH_SIZE = 25

def _update_latest(season_root: Path, run_ts: RunTimestamps, processed_rel: str, excel_rel: Optional[str] = None, stat_map_rel: Optional[str] = None) -> Path:
    """Update _meta/latest.json for the season."""
    meta_dir = season_root / "_meta"
    meta_dir.mkdir(parents=True, exist_ok=True)
    latest_path = meta_dir / "latest.json"

    if latest_path.exists():
        with latest_path.open("r", encoding="utf-8") as f:
            data = json.load(f)
    else:
        data = {"season": season_root.name}

    player_data = data.get("season_player_data")
    if not isinstance(player_data, dict):
        player_data = {}

    player_data["processed"] = str(processed_rel)
    if excel_rel:
        player_data["excel"] = str(excel_rel)
    if stat_map_rel:
        player_data["stat_id_map"] = str(stat_map_rel)

    data["season_player_data"] = player_data
    data["_updated_unix"] = run_ts.unix
    data["_updated_iso_utc"] = run_ts.iso_utc

    with latest_path.open("w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, sort_keys=True)

    return latest_path

def _to_excel(season: str, out_base: Path, stat_map_path: Path, xlsx_path: Path, run_ts: RunTimestamps) -> None:
    """Generate Excel summary of all players generated during this run."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"Players {season}"
    
    # Load stat map
    stat_map = {}
    if stat_map_path.exists():
        stat_map = json.loads(stat_map_path.read_text(encoding="utf-8"))

    # Load all player JSONs
    all_players = []
    for player_dir in out_base.iterdir():
        if player_dir.is_dir():
            for p in player_dir.glob("*.json"):
                try:
                    all_players.append(json.loads(p.read_text(encoding="utf-8")))
                except Exception:
                    pass

    if not all_players:
        ws.append(["No player data found."])
        wb.save(xlsx_path)
        return

    # Determine unique stat IDs across all players
    stat_ids = set()
    for p in all_players:
        for cat in ("season_totals", "advanced_totals"):
            totals = p.get(cat)
            if isinstance(totals, dict):
                stat_ids.update(totals.keys())
    
    sorted_stat_ids = sorted(list(stat_ids), key=lambda x: int(x) if x.isdigit() else x)
    
    # Header
    headers = [
        "player_key", "name_full", "editorial_player_key", "season",
        "uniform_number", "display_position", "primary_position",
        "editorial_team_abbr", "editorial_team_full_name"
    ]
    for sid in sorted_stat_ids:
        headers.append(stat_map.get(sid, f"Stat {sid}"))
    ws.append(headers)

    # Rows
    for p in all_players:
        row = [
            p.get("player_key", ""),
            p.get("name_full", ""),
            p.get("editorial_player_key", ""),
            p.get("season", ""),
            p.get("uniform_number", ""),
            p.get("display_position", ""),
            p.get("primary_position", ""),
            p.get("editorial_team_abbr", ""),
            p.get("editorial_team_full_name", "")
        ]
        
        st = p.get("season_totals") or {}
        at = p.get("advanced_totals") or {}
        for sid in sorted_stat_ids:
            val = st.get(sid)
            if val is None:
                val = at.get(sid, "")
            row.append(str(val))
        ws.append(row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 20

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)




def _dump_json(data: Any, path: Path, pretty: bool) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        if pretty:
            json.dump(data, f, indent=2)
        else:
            json.dump(data, f, separators=(",", ":"))


def _fetch_players_batch(session, player_keys: List[str], season: str, league_key: Optional[str], start: int = 0) -> Dict[str, Any]:
    # If a league_key is provided and player_keys supplied, use the league players batch endpoint.
    # If league_key is None, use the game endpoint for broad player universe: 'game/nhl/players'.
    if league_key and player_keys:
        joined = ",".join(player_keys)
        endpoint = f"league/{league_key}/players;player_keys={joined};out=stats;type=season;season={season}"
    else:
        # Global game endpoint for NHL player universe (broad coverage).
        endpoint = f"game/nhl/players;start={start};count={BATCH_SIZE};out=stats;type=season;season={season}"
    url = f"{API_BASE}/{endpoint}"
    r = session.get(url, params={"format": "json"}, headers={"Accept": "application/json"})
    handle_api_error(r, f"players batch {season}")
    return r.json()


def _flatten_meta_list(meta_list: List[Any]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for item in meta_list:
        if isinstance(item, dict):
            out.update(item)
    return out


def _extract_stats_block(block: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    if not isinstance(block, dict):
        return {}
    stats = block.get("stats") or []
    out: Dict[str, Any] = {}
    for item in stats:
        if not isinstance(item, dict):
            continue
        stat = item.get("stat") or {}
        sid = stat.get("stat_id")
        if sid is None:
            continue
        out[str(sid)] = stat.get("value")
    return out


def _find_game_nodes(obj: Any) -> List[Dict[str, Any]]:
    """Recursively search an object for plausible game nodes."""
    found: List[Dict[str, Any]] = []

    if isinstance(obj, dict):
        # direct game node
        if any(k in obj for k in ("game", "game_id", "game_key", "game_date", "date", "arena")):
            found.append(obj)
        for v in obj.values():
            found.extend(_find_game_nodes(v))
    elif isinstance(obj, list):
        for item in obj:
            found.extend(_find_game_nodes(item))
    return found


def _split_and_write_players(raw_payload: Dict[str, Any], out_base: Path, season: str, run_ts: RunTimestamps, pretty: bool, stat_map: Dict[str, str]) -> None:
    fc = raw_payload.get("fantasy_content") or {}
    # In global/game-scoped payloads, players may exist under league OR game structures.
    container = fc.get("league") or fc.get("game") or fc

    # try to extract players dict via several heuristics
    players_container = None
    if isinstance(container, list) and len(container) >= 2 and isinstance(container[1], dict):
        players_container = container[1].get("players")
    if players_container is None and isinstance(fc.get("players"), dict):
        players_container = fc.get("players")
    if players_container is None:
        # fallback: search fc for any 'players' dict
        players_container = None
        for v in fc.values():
            if isinstance(v, dict) and "players" in v:
                players_container = v.get("players")
                break

    if not isinstance(players_container, dict):
        return

    for k, v in players_container.items():
        if k == "count":
            continue
        player = v.get("player")
        if not isinstance(player, list) or len(player) < 2:
            continue

        meta_list = player[0]
        extra = player[1]
        meta = _flatten_meta_list(meta_list if isinstance(meta_list, list) else [])
        player_key = meta.get("player_key") or str(meta.get("player_id") or "unknown")

        name_info = meta.get("name") or {}
        full_name = name_info.get("full") or f"{name_info.get('first','')} {name_info.get('last','')}".strip()

        # Extract extra metadata
        uniform_number = meta.get("uniform_number")
        display_position = meta.get("display_position")
        primary_position = meta.get("primary_position")
        editorial_team_abbr = meta.get("editorial_team_abbr")
        editorial_team_full_name = meta.get("editorial_team_full_name")
        eligible_positions = meta.get("eligible_positions") or []
        positions = []
        if isinstance(eligible_positions, list):
            for item in eligible_positions:
                if isinstance(item, dict) and item.get("position"):
                    positions.append(item.get("position"))
        
        season_totals = _extract_stats_block(extra.get("player_stats"))
        advanced_totals = _extract_stats_block(extra.get("player_advanced_stats"))

        # Recursively find any game-like nodes in the player's extra payload
        game_nodes = _find_game_nodes(extra)

        out_obj: Dict[str, Any] = {
            "player_key": player_key,
            "player_id": meta.get("player_id"),
            "editorial_player_key": meta.get("editorial_player_key"),
            "name_full": full_name,
            "season": season,
            "uniform_number": uniform_number,
            "display_position": display_position,
            "primary_position": primary_position,
            "editorial_team_abbr": editorial_team_abbr,
            "editorial_team_full_name": editorial_team_full_name,
            "positions": positions,
            "season_totals": season_totals,
            "advanced_totals": advanced_totals,
            "game_entries": [],
            "_generated_unix": run_ts.unix,
            "_generated_iso_utc": run_ts.iso_utc,
        }

        for g in game_nodes:
            if not isinstance(g, dict):
                continue
            # attempt to extract or synthesize a game_id
            game_id = None
            for key in ("game_id", "game_key", "id"):
                if key in g:
                    game_id = str(g.get(key))
                    break
            if not game_id:
                dt = g.get("date") or g.get("game_date") or g.get("time")
                t1 = g.get("home_team") or g.get("team_home") or g.get("home") or g.get("home_team_key")
                t2 = g.get("away_team") or g.get("team_away") or g.get("away") or g.get("away_team_key")
                game_id = f"{dt}|{t1}|{t2}"

            # extract statline from this node if present
            statline = {}
            # possible shapes: g.get('stats') -> list of {stat:{stat_id, value, name}}
            for s_item in (g.get("stats") or []):
                if isinstance(s_item, dict):
                    stat = s_item.get("stat")
                    if isinstance(stat, dict):
                        sid = stat.get("stat_id")
                        if sid is not None:
                            statline[str(sid)] = stat.get("value")
                            name = stat.get("name") or stat.get("display_name")
                            if name:
                                stat_map.setdefault(str(sid), name)

            out_obj["game_entries"].append({"game_id": game_id, "game_raw": g, "statline": statline})

        # Write/merge per-player file
        player_dir = out_base / player_key
        player_dir.mkdir(parents=True, exist_ok=True)
        out_path = player_dir / f"{player_key}.json"
        if out_path.exists():
            try:
                existing = json.loads(out_path.read_text(encoding="utf-8"))
                existing["season_totals"] = out_obj.get("season_totals") or existing.get("season_totals")
                existing["advanced_totals"] = out_obj.get("advanced_totals") or existing.get("advanced_totals")
                existing["_generated_unix"] = out_obj.get("_generated_unix")
                existing["_generated_iso_utc"] = out_obj.get("_generated_iso_utc")

                existing_game_ids = {ge.get("game_id") for ge in existing.get("game_entries", []) if isinstance(ge, dict)}
                for ge in out_obj.get("game_entries", []):
                    gid = ge.get("game_id")
                    if gid not in existing_game_ids:
                        existing.setdefault("game_entries", []).append(ge)
                _dump_json(existing, out_path, pretty)
            except Exception:
                _dump_json(out_obj, out_path, pretty)
        else:
            _dump_json(out_obj, out_path, pretty)


def main() -> None:
    ap = argparse.ArgumentParser(description="Fetch per-player season data and write per-player JSON files.")
    ap.add_argument("--league-key", required=False, help="Optional full league key (if provided, will use rostered players). If omitted the script will fetch the global game player universe via the game endpoint.")
    ap.add_argument("--season", required=True, help="Season year, e.g. 2025")
    ap.add_argument("--stale-hours", type=float, default=0.0, help="Not used (no cache) but kept for parity")
    ap.add_argument("--pretty", action="store_true")
    ap.add_argument("--to-excel", action="store_true", help="Also write an Excel workbook summary")
    args = ap.parse_args()

    season = str(args.season)
    run_ts = make_run_timestamps()

    # Determine player universe: prefer rostered players when league_key provided,
    # otherwise fetch the broader game-scoped player universe via the game endpoint
    session = get_session()
    season_root = get_export_dir() / season
    out_base = season_root / "playerdata"
    stat_map: Dict[str, str] = {}
    
    # Query Yahoo global stat_categories API to seed the exact valid stat_map IDs and names
    try:
        stat_cats_url = f"{API_BASE}/game/{args.game if hasattr(args, 'game') else 'nhl'}/stat_categories?format=json"
        cat_r = session.get(stat_cats_url, headers={"Accept": "application/json"})
        cat_j = cat_r.json()
        cat_fc = cat_j.get("fantasy_content", {})
        cat_game = cat_fc.get("game", [])
        if len(cat_game) >= 2 and isinstance(cat_game[1], dict):
            stats_list = cat_game[1].get("stat_categories", {}).get("stats", [])
            for st_entry in stats_list:
                if isinstance(st_entry, dict):
                    stat = st_entry.get("stat", {})
                    sid = stat.get("stat_id")
                    name = stat.get("name") or stat.get("display_name")
                    if sid is not None and name:
                        stat_map[str(sid)] = str(name)
    except Exception as e:
        print(f"Warning: could not seed global stat_categories map. {e}", file=sys.stderr)

    player_keys: List[str] = []
    if args.league_key:
        league_root = get_export_dir() / args.league_key
        latest_p = league_root / "_meta" / "latest.json"
        if not latest_p.exists():
            print("ERROR: missing _meta/latest.json for league; run league_dump + rostered list first", file=sys.stderr)
            sys.exit(1)
        latest = json.loads(latest_p.read_text(encoding="utf-8"))
        roster_block = latest.get("league_rostered_players_list")
        if not roster_block or "processed" not in roster_block:
            print("ERROR: league_rostered_players_list not found in latest.json", file=sys.stderr)
            sys.exit(1)
        roster_path = league_root / roster_block["processed"]
        roster = json.loads(roster_path.read_text(encoding="utf-8"))
        player_items = roster.get("players") or []
        player_keys = [p.get("player_key") for p in player_items if isinstance(p, dict) and p.get("player_key")]
        player_keys = sorted(set(player_keys))
        if not player_keys:
            print("No rostered players found; nothing to do.")
            sys.exit(0)

        total = len(player_keys)
        # 2-hour cache check (7200 seconds)
        # Skip this check if no league-key is provided (global run)
        cache_ttl = 7200
        players_to_fetch = []
        for pk in player_keys:
            p_path = out_base / pk / f"{pk}.json"
            if p_path.exists():
                try:
                    p_data = json.loads(p_path.read_text(encoding="utf-8"))
                    gen_unix = p_data.get("_generated_unix")
                    if gen_unix and (run_ts.unix - gen_unix) < cache_ttl:
                        continue # Skip fetch, cache is fresh
                except Exception:
                    pass
            players_to_fetch.append(pk)

        if players_to_fetch:
            print(f"Fetching {len(players_to_fetch)} players (skipped {len(player_keys) - len(players_to_fetch)} fresh cache entries)")
            total_fetch = len(players_to_fetch)
            batches = math.ceil(total_fetch / BATCH_SIZE)
            for i in range(batches):
                batch = players_to_fetch[i * BATCH_SIZE : (i + 1) * BATCH_SIZE]
                raw = _fetch_players_batch(session, batch, season, args.league_key)
                _split_and_write_players(raw, out_base, season, run_ts, args.pretty, stat_map)
        else:
            print("All players in league have fresh cache (< 2 hours old). Skipping API fetches.")
    else:
        # No league key: fetch the global game player universe for the season
        # Using the game endpoint with 'nhl' game_key to request broad player coverage.
        # We must paginate through the results using start and count.
        start = 0
        total_processed = 0
        while True:
            raw = _fetch_players_batch(session, [], season, None, start=start)
            
            # Check if we got any players back to determine if we should stop
            fc = raw.get("fantasy_content") or {}
            container = fc.get("game") or []
            players_count = 0
            
            if isinstance(container, list) and len(container) >= 2 and isinstance(container[1], dict):
                players_container = container[1].get("players")
                if isinstance(players_container, dict):
                    players_count = players_container.get("count", 0)
            
            if not players_count or int(players_count) == 0:
                print(f"No more players found (start={start}). Ending pagination.")
                break
                
            _split_and_write_players(raw, out_base, season, run_ts, args.pretty, stat_map)
            
            print(f"Fetched and processed {players_count} players (start={start})")
            total_processed += int(players_count)
            
            if int(players_count) < BATCH_SIZE:
                break
                
            start += BATCH_SIZE
            
        print(f"Total global players processed: {total_processed}")

    # Merge with any existing stat_map on disk (keep prior names)
    existing_map_glob = out_base.glob("stat_id_map.*.json")
    for p in existing_map_glob:
        try:
            m = json.loads(p.read_text(encoding="utf-8"))
            stat_map.update({k: v for k, v in m.items() if k not in stat_map})
        except Exception:
            continue

    # write stat id map
    stat_map_path = out_base / f"stat_id_map.{run_ts.iso_stamp}.json"
    _dump_json(stat_map, stat_map_path, args.pretty)
    stat_map_rel = stat_map_path.relative_to(season_root).as_posix()
    print(f"Wrote stat id map: {stat_map_path}")

    processed_rel = out_base.relative_to(season_root).as_posix()
    excel_rel = None

    if args.to_excel:
        excel_path = season_root / "playerdata" / f"player_summary.{run_ts.iso_stamp}.xlsx"
        _to_excel(season, out_base, stat_map_path, excel_path, run_ts)
        excel_rel = excel_path.relative_to(season_root).as_posix()
        print(f"Wrote Excel summary: {excel_path}")

    latest_path = _update_latest(season_root, run_ts, processed_rel, excel_rel, stat_map_rel)
    print(f"Updated latest.json: {latest_path}")


if __name__ == "__main__":
    main()
