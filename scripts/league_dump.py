#!/usr/bin/env python3
from __future__ import annotations
"""
League dump: metadata + teams + scoring settings (no standings/transactions).

Outputs:
  - league_info.json
  - league_teams.json
  - league_scoring.json
  - (optional) league_info.xlsx (polished formatting + team logos)

Depends on:
  - src/auth/oauth.get_session()
  - src/config/env.get_export_dir()
"""

import argparse, io, json
from pathlib import Path
from typing import Any, Dict, List, Optional
import requests

from src.auth.oauth import get_session
from src.config.env import get_export_dir

BASE_URL = "https://fantasysports.yahooapis.com/fantasy/v2"


# ---------------- HTTP ----------------
def _fetch(endpoint: str) -> dict:
    sess = get_session()
    url = f"{BASE_URL}/{endpoint}"
    r = sess.get(url, params={"format": "json"}, headers={"Accept": "application/json"})
    r.raise_for_status()
    return r.json()


# ---------------- Extractors ----------------
_META_KEYS = {
    "league_key","league_id","name","url","logo_url","draft_status","num_teams",
    "edit_key","weekly_deadline","roster_type","league_update_timestamp","scoring_type",
    "league_type","renew","renewed","felo_tier","matchup_week",
    "short_invitation_url","allow_add_to_dl_extra_pos","is_pro_league","is_cash_league",
    "current_week","start_week","start_date","end_week","end_date","is_finished","current_date",
    "game_code","season"
}

def _extract_league_info(payload: dict) -> dict:
    fc = payload.get("fantasy_content", {})
    league = fc.get("league")
    info: dict = {}
    if isinstance(league, list):
        for entry in league:
            if isinstance(entry, dict):
                for k in _META_KEYS:
                    if k in entry:
                        info[k] = entry[k]
                if info:
                    break
    elif isinstance(league, dict):
        for k in _META_KEYS:
            if k in league:
                info[k] = league[k]
    return info


def _flatten_singleton_dicts(node) -> dict:
    flat: dict = {}
    if isinstance(node, list):
        for item in node:
            if isinstance(item, dict):
                flat.update(item)
    elif isinstance(node, dict):
        flat.update(node)
    return flat


def _extract_teams(payload: dict) -> List[dict]:
    fc = payload.get("fantasy_content", {})
    league = fc.get("league")
    teams_out: List[dict] = []
    teams_container = None
    if isinstance(league, list):
        for entry in league:
            if isinstance(entry, dict) and "teams" in entry:
                teams_container = entry["teams"]
                break
    elif isinstance(league, dict) and "teams" in league:
        teams_container = league["teams"]

    if isinstance(teams_container, dict):
        for v in teams_container.values():
            if isinstance(v, dict) and "team" in v:
                node = v["team"]
                team_arr = node[0] if isinstance(node, list) else node
                flat = _flatten_singleton_dicts(team_arr)

                # First manager (if present)
                manager = None
                if isinstance(node, list):
                    for item in node:
                        if isinstance(item, dict) and "managers" in item:
                            mgrs = item["managers"]
                            if isinstance(mgrs, list) and mgrs:
                                # common shape: [{"manager": {...}}]
                                m0 = mgrs[0]
                                manager = m0.get("manager") if isinstance(m0, dict) else m0
                                break

                # logo url
                logo_url = None
                logos = flat.get("team_logos")
                if isinstance(logos, list) and logos and isinstance(logos[0], dict):
                    tl = logos[0].get("team_logo")
                    if isinstance(tl, dict):
                        logo_url = tl.get("url")

                teams_out.append({
                    "team_key": flat.get("team_key"),
                    "team_id": flat.get("team_id"),
                    "name": flat.get("name"),
                    "url": flat.get("url"),
                    "logo": logo_url,
                    "waiver_priority": flat.get("waiver_priority"),
                    "moves": flat.get("number_of_moves"),
                    "trades": flat.get("number_of_trades"),
                    "manager": manager,
                })

    return teams_out


def _extract_settings(payload: dict) -> dict:
    """
    Produces:
      - stat_categories: [{stat_id, display_name, name, position_type, group, abbr, is_only_display_stat}]
      - stat_modifiers:  [{stat_id, value}]
      - roster_positions:[{position, position_type, count, is_starting_position}]
      - tiebreakers:     [str or dict] (best-effort)
      - head_to_head:    {draft_type, uses_playoff, waiver_type, ...}
      - goalie_minimums: {min_games_played, week_has_enough_qualifying_days: {...}}  (if present)
    """
    fc = payload.get("fantasy_content", {})
    league = fc.get("league")
    settings = None

    if isinstance(league, list):
        for entry in league:
            if isinstance(entry, dict) and "settings" in entry:
                settings = entry["settings"]
                break
    elif isinstance(league, dict):
        settings = league.get("settings")

    base_cfg = {}
    addl_cfg = {}
    if isinstance(settings, list):
        if len(settings) > 0 and isinstance(settings[0], dict):
            base_cfg = settings[0]
        if len(settings) > 1 and isinstance(settings[1], dict):
            addl_cfg = settings[1]
    elif isinstance(settings, dict):
        base_cfg = settings

    # roster positions
    roster_positions: List[dict] = []
    rp = base_cfg.get("roster_positions")
    if isinstance(rp, list):
        for item in rp:
            if isinstance(item, dict) and "roster_position" in item:
                node = item["roster_position"]
                if isinstance(node, dict):
                    roster_positions.append({
                        "position": node.get("position"),
                        "position_type": node.get("position_type"),
                        "count": node.get("count"),
                        "is_starting_position": node.get("is_starting_position"),
                    })

    # stat categories
    stat_categories: List[dict] = []
    sc = base_cfg.get("stat_categories", {})
    stats = sc.get("stats") if isinstance(sc, dict) else None
    if isinstance(stats, list):
        for s in stats:
            if isinstance(s, dict) and "stat" in s and isinstance(s["stat"], dict):
                st = s["stat"]
                stat_categories.append({
                    "stat_id": st.get("stat_id"),
                    "name": st.get("name"),
                    "display_name": st.get("display_name"),
                    "group": st.get("group"),
                    "abbr": st.get("abbr"),
                    "position_type": st.get("position_type"),
                    "is_only_display_stat": st.get("is_only_display_stat"),
                })

    # stat modifiers
    stat_modifiers: List[dict] = []
    sm = base_cfg.get("stat_modifiers")
    if isinstance(sm, dict):
        sm_stats = sm.get("stats")
        if isinstance(sm_stats, list):
            for entry in sm_stats:
                if isinstance(entry, dict) and "stat" in entry and isinstance(entry["stat"], dict):
                    st = entry["stat"]
                    stat_modifiers.append({"stat_id": st.get("stat_id"), "value": st.get("value")})

    # tiebreakers
    tiebreakers: List = []
    for key in ("tiebreakers", "tiebreaker_rules", "playoff_tie_breaker_rules"):
        tb = base_cfg.get(key)
        if tb:
            if isinstance(tb, list):
                tiebreakers = tb
            elif isinstance(tb, dict):
                for v in tb.values():
                    if isinstance(v, str):
                        tiebreakers.append(v)
                    elif isinstance(v, dict):
                        tiebreakers.append(v.get("rule") or v.get("name") or v)
            break

    # head-to-head keys
    h2h_keys = [
        "draft_type","is_auction_draft","scoring_type","invite_permission",
        "uses_playoff","has_playoff_consolation_games","playoff_start_week",
        "uses_playoff_reseeding","uses_lock_eliminated_teams","num_playoff_teams",
        "num_playoff_consolation_teams","has_multiweek_championship","waiver_type",
        "waiver_rule","uses_faab","draft_time","draft_pick_time","post_draft_players",
        "max_teams","waiver_time","trade_end_date","trade_ratify_type","trade_reject_time",
        "player_pool","cant_cut_list","draft_together","can_trade_draft_picks",
        "max_weekly_adds","uses_median_score"
    ]
    head_to_head = {k: base_cfg.get(k) for k in h2h_keys if k in base_cfg}

    # goalie minimums (if present)
    goalie_minimums = {}
    if isinstance(addl_cfg, dict):
        if "min_games_played" in addl_cfg:
            goalie_minimums["min_games_played"] = addl_cfg["min_games_played"]
        if "week_has_enough_qualifying_days" in addl_cfg:
            goalie_minimums["week_has_enough_qualifying_days"] = addl_cfg["week_has_enough_qualifying_days"]

    return {
        "stat_categories": stat_categories,
        "stat_modifiers": stat_modifiers,
        "roster_positions": roster_positions,
        "tiebreakers": tiebreakers,
        "head_to_head": head_to_head,
        "goalie_minimums": goalie_minimums,
    }


# ---------------- IO ----------------
def _dump_json(obj, filename: str, pretty: bool) -> Path:
    outdir = get_export_dir()
    outdir.mkdir(parents=True, exist_ok=True)
    path = outdir / filename
    with path.open("w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2 if pretty else None)
    print(f"Wrote {path}")
    return path


# ---------------- Excel ----------------
def _to_excel(league_info: dict, teams: List[dict], scoring: dict, xlsx_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font
    from openpyxl.drawing.image import Image as XLImage

    session = get_session()
    wb = Workbook()

    # League sheet
    ws = wb.active
    ws.title = "League"
    ws.append(("Field", "Value"))
    for k, v in league_info.items():
        ws.append((k, v))
        if k == "url" and isinstance(v, str) and v.startswith("http"):
            cell = ws.cell(row=ws.max_row, column=2)
            cell.hyperlink = v
            cell.font = Font(underline="single", color="0000EE")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:B{ws.max_row}"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80

    # Teams sheet
    ws2 = wb.create_sheet("Teams")
    headers = ["team_id", "name", "url", "logo", "waiver_priority", "moves", "trades"]
    ws2.append(headers)

    def _download_image(url: str) -> Optional[bytes]:
        if not url or not isinstance(url, str) or not url.startswith("http"):
            return None
        try:
            r = session.get(url, timeout=20)
            r.raise_for_status()
            return r.content
        except Exception:
            return None

    for t in teams:
        ws2.append([
            t.get("team_id"),
            t.get("name"),
            t.get("url"),
            "",  # placeholder for logo image
            t.get("waiver_priority"),
            t.get("moves"),
            t.get("trades"),
        ])
        r_idx = ws2.max_row
        # hyperlink on team name
        name_cell = ws2.cell(row=r_idx, column=2)
        url_val = t.get("url")
        if isinstance(url_val, str) and url_val.startswith("http"):
            name_cell.hyperlink = url_val
            name_cell.font = Font(underline="single", color="0000EE")

        # embed logo at col D
        logo_bytes = _download_image(t.get("logo"))
        if logo_bytes:
            img = XLImage(io.BytesIO(logo_bytes))
            img.width, img.height = 32, 32
            ws2.add_image(img, f"D{r_idx}")
            ws2.row_dimensions[r_idx].height = 26

    ws2.freeze_panes = "A2"
    widths = {"team_id":12,"name":30,"url":60,"logo":8,"waiver_priority":16,"moves":10,"trades":10}
    for i, col in enumerate(headers, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = widths.get(col, 14)
    ws2.auto_filter.ref = f"A1:{get_column_letter(ws2.max_column)}{ws2.max_row}"

    # Scoring sheets
    def write_table(sheet_name: str, rows: List[dict] | List[Any], headers_hint: List[str] | None = None):
        wsx = wb.create_sheet(sheet_name)
        if not rows:
            wsx.append(["(empty)"])
            return
        if isinstance(rows[0], dict):
            keys = headers_hint or sorted({k for r in rows for k in r.keys()})
            wsx.append(keys)
            for r in rows:
                wsx.append([r.get(k) for k in keys])
            wsx.freeze_panes = "A2"
            wsx.auto_filter.ref = f"A1:{get_column_letter(wsx.max_column)}{wsx.max_row}"
            for c in range(1, wsx.max_column + 1):
                wsx.column_dimensions[get_column_letter(c)].width = 18
        else:
            wsx.append(["value"])
            for v in rows:
                wsx.append([v])
            wsx.freeze_panes = "A2"
            wsx.auto_filter.ref = f"A1:A{wsx.max_row}"
            wsx.column_dimensions["A"].width = 40

    write_table("ScoringCategories", scoring.get("stat_categories", []),
                headers_hint=["stat_id","display_name","name","group","abbr","position_type","is_only_display_stat"])
    write_table("StatModifiers", scoring.get("stat_modifiers", []), headers_hint=["stat_id","value"])
    write_table("RosterPositions", scoring.get("roster_positions", []),
                headers_hint=["position","position_type","count","is_starting_position"])

    # Head-to-head knobs + goalie minimums
    h2h_items = [{"key": k, "value": v} for k, v in (scoring.get("head_to_head") or {}).items()]
    write_table("HeadToHeadSettings", h2h_items, headers_hint=["key","value"])
    gm = scoring.get("goalie_minimums") or {}
    gm_rows = []
    if "min_games_played" in gm:
        gm_rows.append({"key": "min_games_played", "value": gm.get("min_games_played")})
    weeks = gm.get("week_has_enough_qualifying_days")
    if isinstance(weeks, dict):
        for wk, ok in weeks.items():
            gm_rows.append({"key": f"week_{wk}", "value": ok})
    write_table("GoalieMinimums", gm_rows, headers_hint=["key","value"])

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


# ---------------- CLI ----------------
def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Dump Yahoo Fantasy league info (metadata + teams + scoring).")
    g = p.add_mutually_exclusive_group(required=True)
    g.add_argument("--league-key", help="Full league key, e.g., 453.l.33099")
    g.add_argument("--league-id", type=int, help="League ID (use with --game)")
    p.add_argument("--game", default="nhl", help="Game key (default: nhl)")
    p.add_argument("--pretty", action="store_true", help="Pretty-print JSON")
    p.add_argument("--to-excel", action="store_true", help="Also write an Excel workbook")
    return p.parse_args()


def _resolve_league_key(args: argparse.Namespace) -> str:
    if args.league_key:
        return args.league_key
    return f"{args.game}.l.{args.league_id}"


def main() -> None:
    args = _parse_args()
    league_key = _resolve_league_key(args)

    # Fetch raw
    meta_raw = _fetch(f"league/{league_key}/metadata")
    teams_raw = _fetch(f"league/{league_key}/teams")
    settings_raw = _fetch(f"league/{league_key}/settings")

    # Extract
    league_info = _extract_league_info(meta_raw)
    teams = _extract_teams(teams_raw)
    scoring = _extract_settings(settings_raw)

    # Write JSON
    _dump_json(league_info, "league_info.json", args.pretty)
    _dump_json(teams, "league_teams.json", args.pretty)
    _dump_json(scoring, "league_scoring.json", args.pretty)

    # Optional Excel
    if args.to_excel:
        xlsx_path = get_export_dir() / "league_info.xlsx"
        _to_excel(league_info, teams, scoring, xlsx_path)
        print(f"Wrote Excel: {xlsx_path}")


if __name__ == "__main__":
    main()
