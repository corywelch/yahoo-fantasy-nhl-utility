#!/usr/bin/env python3
from __future__ import annotations

"""
standings_dump: week-by-week standings reconstruction for a Yahoo Fantasy league.

This script depends on a prior run of ``league_dump`` for the same league.
It *reads* the latest processed league dump JSON from:

    exports/<league_key>/_meta/latest.json  →  league_dump.processed

and then fetches per-week scoreboard data to build:

  exports/<league_key>/standings_dump/
    raw/
      scoreboard.wkNNN.<ISO>.json
    processed/
      matchups.<ISO>.json
      weekly.<ISO>.json
      summary.<ISO>.json
    excel/
      standings.wkSS-EE.<ISO>.xlsx
    manifest/
      manifest.<ISO>.json

Where:
  - <league_key> is the full Yahoo league key, e.g. "nhl.453.l.33099"
  - <ISO> is a run identifier like "20251129T014755Z"
  - SS / EE are 2-digit start/end week numbers (e.g. 02-23).

It also updates:

  exports/<league_key>/_meta/latest.json

by adding/updating a "standings_dump" block pointing at the latest
processed JSON and Excel paths (relative to the league root).
"""

import argparse
import json
import hashlib
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from src.auth.oauth import get_session
from src.config.env import get_export_dir
from src.util_time import make_run_timestamps, RunTimestamps

BASE_URL = "https://fantasysports.yahooapis.com/fantasy/v2"


# ---------------- HTTP helpers ----------------


def _fetch(endpoint: str) -> dict:
    """
    Fetch a Fantasy API endpoint as JSON, raising for HTTP errors.
    """
    sess = get_session()
    url = f"{BASE_URL}/{endpoint}"
    r = sess.get(url, params={"format": "json"}, headers={"Accept": "application/json"})
    r.raise_for_status()
    return r.json()


def _fetch_scoreboard_week(league_key: str, week: int) -> dict:
    """
    Fetch /league/{league_key}/scoreboard;week={week} as JSON.
    """
    endpoint = f"league/{league_key}/scoreboard;week={week}"
    return _fetch(endpoint)


# ---------------- league_dump context loading ----------------


@dataclass
class LeagueContext:
    league_key: str
    league_info: Dict[str, Any]
    teams: List[Dict[str, Any]]
    scoring: Dict[str, Any]


def _load_latest_league_dump(league_key: str) -> LeagueContext:
    """
    Load the latest processed league_dump JSON for this league from _meta/latest.json.

    Error and exit if no league_dump entry is present, since standings_dump
    is designed to build on top of league_dump outputs.
    """
    root = get_export_dir() / league_key
    meta_dir = root / "_meta"
    latest_path = meta_dir / "latest.json"

    if not latest_path.exists():
        print(f"[ERROR] No _meta/latest.json for league {league_key}.")
        print("        Run league_dump first, then re-run standings_dump.")
        raise SystemExit(1)

    with latest_path.open("r", encoding="utf-8") as f:
        latest = json.load(f)

    ld = latest.get("league_dump")
    if not ld or "processed" not in ld:
        print(f"[ERROR] _meta/latest.json for {league_key} has no 'league_dump.processed' entry.")
        print("        Run league_dump first, then re-run standings_dump.")
        raise SystemExit(1)

    rel_processed = ld["processed"]
    processed_path = root / rel_processed

    if not processed_path.exists():
        print(f"[ERROR] league_dump processed file not found: {processed_path}")
        print("        Re-run league_dump for this league, then re-run standings_dump.")
        raise SystemExit(1)

    with processed_path.open("r", encoding="utf-8") as f:
        payload = json.load(f)

    league_info = payload.get("league_info") or {}
    teams = payload.get("teams") or []
    scoring = payload.get("scoring") or {}

    return LeagueContext(
        league_key=league_key,
        league_info=league_info,
        teams=teams,
        scoring=scoring,
    )


# ---------------- paths + manifest helpers ----------------


@dataclass
class StandingsPaths:
    league_key: str
    root: Path
    meta_dir: Path
    raw_dir: Path
    processed_dir: Path
    excel_dir: Path
    manifest_dir: Path


def _prepare_standings_dirs(league_key: str) -> StandingsPaths:
    """
    Prepare league-scoped export directories under exports/<league_key>/standings_dump/...
    """
    base = get_export_dir()
    league_root = base / league_key
    meta_dir = league_root / "_meta"
    sd_root = league_root / "standings_dump"
    raw_dir = sd_root / "raw"
    processed_dir = sd_root / "processed"
    excel_dir = sd_root / "excel"
    manifest_dir = sd_root / "manifest"

    for d in (meta_dir, raw_dir, processed_dir, excel_dir, manifest_dir):
        d.mkdir(parents=True, exist_ok=True)

    return StandingsPaths(
        league_key=league_key,
        root=league_root,
        meta_dir=meta_dir,
        raw_dir=raw_dir,
        processed_dir=processed_dir,
        excel_dir=excel_dir,
        manifest_dir=manifest_dir,
    )


def _sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _build_manifest_dict(
    module_name: str,
    league_key: str,
    paths: StandingsPaths,
    run_ts: RunTimestamps,
    cli_args: Dict[str, Any],
    produced_paths: List[Path],
) -> Dict[str, Any]:
    files: Dict[str, Dict[str, Any]] = {}

    for abs_path in produced_paths:
        # Normalize to league-root-relative POSIX paths
        rel = abs_path.relative_to(paths.root).as_posix()
        stat = abs_path.stat()
        files[rel] = {
            "size_bytes": stat.st_size,
            "sha256": _sha256_file(abs_path),
        }

    return {
        "module": module_name,
        "league_key": league_key,
        "_generated_unix": run_ts.unix,
        "_generated_iso_utc": run_ts.iso_utc,
        "_generated_iso_local": run_ts.iso_local,
        "files": files,
        "cli_args": cli_args,
    }


def _write_manifest(paths: StandingsPaths, run_ts: RunTimestamps, manifest_data: Dict[str, Any]) -> Path:
    out_path = paths.manifest_dir / f"manifest.{run_ts.iso_stamp}.json"
    with out_path.open("w", encoding="utf-8") as f:
        json.dump(manifest_data, f, indent=2, sort_keys=True)
    return out_path


def _update_latest(
    paths: StandingsPaths,
    run_ts: RunTimestamps,
    matchups_rel: str,
    weekly_rel: str,
    summary_rel: str,
    excel_rel: Optional[str],
) -> Path:
    """
    Update _meta/latest.json with a 'standings_dump' block, preserving other modules' keys.
    """
    latest_path = paths.meta_dir / "latest.json"

    if latest_path.exists():
        with latest_path.open("r", encoding="utf-8") as f:
            data = json.load(f)
    else:
        data = {"league_key": paths.league_key}

    sd = data.get("standings_dump", {})
    sd["matchups"] = matchups_rel
    sd["weekly"] = weekly_rel
    sd["summary"] = summary_rel
    if excel_rel is not None:
        sd["excel"] = excel_rel

    data["standings_dump"] = sd
    data["_updated_unix"] = run_ts.unix
    data["_updated_iso_utc"] = run_ts.iso_utc

    with latest_path.open("w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, sort_keys=True)

    return latest_path


# ---------------- scoreboard parsing ----------------


def _get_scoreboard_node(payload: dict) -> Optional[dict]:
    """
    Given a scoreboard payload, return the "scoreboard" node or None if absent.
    """
    fc = payload.get("fantasy_content") or {}
    league_node = fc.get("league")

    if isinstance(league_node, list):
        for entry in league_node:
            if isinstance(entry, dict) and "scoreboard" in entry:
                return entry["scoreboard"]
    elif isinstance(league_node, dict) and "scoreboard" in league_node:
        return league_node["scoreboard"]

    return None


def _iter_matchups(scoreboard_node: dict) -> Iterable[dict]:
    """
    Yield matchup dicts from a scoreboard node.
    """
    if not scoreboard_node:
        return

    container = None
    if "0" in scoreboard_node:
        # usual shape: {"week": "N", "0": {"matchups": {...}}}
        node0 = scoreboard_node["0"]
        if isinstance(node0, dict):
            container = node0.get("matchups")

    if not isinstance(container, dict):
        return

    count = int(container.get("count", 0))
    for i in range(count):
        mwrap = container.get(str(i))
        if not (isinstance(mwrap, dict) and "matchup" in mwrap):
            continue
        matchup = mwrap["matchup"]
        if isinstance(matchup, dict):
            yield matchup


def _flatten_team_core_list(core_list: List[Any]) -> Dict[str, Any]:
    """
    Given the first element of a Yahoo "team" array (list of singleton dicts),
    flatten it into a single dict.
    """
    flat: Dict[str, Any] = {}
    for item in core_list:
        if isinstance(item, dict):
            flat.update(item)
    return flat


def _parse_team_node(team_node: List[Any]) -> Tuple[str, Dict[str, Any], Dict[str, float], Optional[float]]:
    """
    Parse a Yahoo scoreboard "team" array into:

      (team_key, team_meta, stats_by_stat_id, team_points_total)
    """
    if not team_node or not isinstance(team_node, list):
        return "", {}, {}, None

    core_list = team_node[0]
    flat = _flatten_team_core_list(core_list)

    team_key = str(flat.get("team_key") or "")
    team_meta = {
        "team_key": team_key,
        "team_id": flat.get("team_id"),
        "name": flat.get("name"),
        "url": flat.get("url"),
    }

    team_stats_block = None
    team_points_block = None
    for item in team_node[1:]:
        if not isinstance(item, dict):
            continue
        if "team_stats" in item:
            team_stats_block = item["team_stats"]
        if "team_points" in item:
            team_points_block = item["team_points"]

    stats_by_id: Dict[str, float] = {}
    if isinstance(team_stats_block, dict):
        for s in team_stats_block.get("stats", []):
            if not isinstance(s, dict):
                continue
            st = s.get("stat") or {}
            stat_id = str(st.get("stat_id"))
            value = st.get("value")
            if value is None or stat_id == "None":
                continue
            try:
                stats_by_id[stat_id] = float(value)
            except (TypeError, ValueError):
                # keep as best-effort; don't crash on weird values
                continue

    team_points_total: Optional[float] = None
    if isinstance(team_points_block, dict):
        total = team_points_block.get("total")
        try:
            team_points_total = float(total)
        except (TypeError, ValueError):
            team_points_total = None

    return team_key, team_meta, stats_by_id, team_points_total


def _build_stat_winners_map(matchup: dict) -> Dict[str, Dict[str, Any]]:
    """
    Build a map: stat_id -> {"winner_team_key": str | None, "is_tied": bool}.
    """
    out: Dict[str, Dict[str, Any]] = {}
    for entry in matchup.get("stat_winners", []):
        if not isinstance(entry, dict):
            continue
        sw = entry.get("stat_winner") or {}
        stat_id = str(sw.get("stat_id"))
        if not stat_id or stat_id == "None":
            continue
        is_tied = bool(sw.get("is_tied"))
        winner_team_key = sw.get("winner_team_key")
        out[stat_id] = {
            "winner_team_key": str(winner_team_key) if winner_team_key else None,
            "is_tied": is_tied,
        }
    return out


def _compute_weekly_results_for_matchup(
    matchup: dict,
    scoring_stat_ids: List[str],
) -> Tuple[
    Dict[str, Any],  # matchup summary
    List[Dict[str, Any]],  # per-team weekly rows (without prev_opponent)
]:
    """
    Compute per-matchup summary and per-team weekly rows, using scoreboard data.
    """
    week_num = int(matchup.get("week"))
    week_start = matchup.get("week_start")
    week_end = matchup.get("week_end")
    is_playoffs = str(matchup.get("is_playoffs") or "0") == "1"
    is_consolation = str(matchup.get("is_consolation") or "0") == "1"
    is_tied = bool(matchup.get("is_tied"))
    winner_team_key = str(matchup.get("winner_team_key") or "") or None

    teams_container = matchup.get("0", {}).get("teams")
    if not isinstance(teams_container, dict) or "count" not in teams_container:
        return {}, []

    team_entries: List[Dict[str, Any]] = []
    count = int(teams_container.get("count", 0))
    for idx in range(count):
        wrap = teams_container.get(str(idx))
        if not (isinstance(wrap, dict) and "team" in wrap):
            continue
        team_node = wrap["team"]
        if not isinstance(team_node, list):
            continue
        team_key, team_meta, stats_by_id, team_points_total = _parse_team_node(team_node)
        if not team_key:
            continue
        team_entries.append(
            {
                "team_key": team_key,
                "meta": team_meta,
                "stats": stats_by_id,
                "team_points": team_points_total,
            }
        )

    if len(team_entries) != 2:
        # For now, only handle standard 1v1 matchups.
        return {}, []

    team_a = team_entries[0]
    team_b = team_entries[1]

    # Basic matchup summary
    matchup_summary = {
        "week": week_num,
        "week_start": week_start,
        "week_end": week_end,
        "is_playoffs": is_playoffs,
        "is_consolation": is_consolation,
        "winner_team_key": winner_team_key,
        "is_tied": is_tied,
        "team_a_key": team_a["team_key"],
        "team_b_key": team_b["team_key"],
        "team_a_points": team_a.get("team_points"),
        "team_b_points": team_b.get("team_points"),
    }

    # H2H category wins/losses/ties per team, based on stat_winners.
    sw_map = _build_stat_winners_map(matchup)
    results: Dict[str, Dict[str, int]] = {
        team_a["team_key"]: {"wins": 0, "losses": 0, "ties": 0},
        team_b["team_key"]: {"wins": 0, "losses": 0, "ties": 0},
    }

    for stat_id in scoring_stat_ids:
        entry = sw_map.get(stat_id)
        if not entry:
            # Yahoo usually includes all scoring stats in stat_winners, but if not,
            # treat missing as a tie to be safe.
            results[team_a["team_key"]]["ties"] += 1
            results[team_b["team_key"]]["ties"] += 1
            continue

        if entry.get("is_tied"):
            results[team_a["team_key"]]["ties"] += 1
            results[team_b["team_key"]]["ties"] += 1
            continue

        winner_key = entry.get("winner_team_key")
        if not winner_key or winner_key not in results:
            # Unknown winner; skip this stat rather than crashing.
            continue

        loser_key = team_b["team_key"] if winner_key == team_a["team_key"] else team_a["team_key"]
        results[winner_key]["wins"] += 1
        results[loser_key]["losses"] += 1

    # Build weekly rows (prev_opponent filled later).
    weekly_rows: List[Dict[str, Any]] = []
    for home, away in ((team_a, team_b), (team_b, team_a)):
        weekly_rows.append(
            {
                "week": week_num,
                "team_key": home["team_key"],
                "opponent_key": away["team_key"],
                "prev_opponent_key": None,  # filled after all weeks are loaded
                "is_playoffs": is_playoffs,
                "is_consolation": is_consolation,
                "categories": home["stats"],  # {stat_id: value}
                "result": results[home["team_key"]],
                "team_points": home.get("team_points"),
            }
        )

    return matchup_summary, weekly_rows


def _backfill_prev_opponents(weekly_rows: List[Dict[str, Any]]) -> None:
    """
    For each (team, week) row, set prev_opponent_key based on the prior week's opponent.
    """
    by_team: Dict[str, List[Tuple[int, str]]] = {}
    for row in weekly_rows:
        team_key = row.get("team_key")
        week = int(row.get("week"))
        opp = row.get("opponent_key")
        if not team_key:
            continue
        by_team.setdefault(team_key, []).append((week, opp))

    # For each team, sort by week and compute previous opponent.
    prev_map: Dict[Tuple[str, int], Optional[str]] = {}
    for team_key, items in by_team.items():
        items.sort(key=lambda x: x[0])
        prev_opp: Optional[str] = None
        for week, opp in items:
            prev_map[(team_key, week)] = prev_opp
            prev_opp = opp

    # Apply back to rows
    for row in weekly_rows:
        team_key = row.get("team_key")
        week = int(row.get("week"))
        row["prev_opponent_key"] = prev_map.get((team_key, week))


# ---------------- summary aggregation ----------------


def _aggregate_summary(
    league_key: str,
    weekly_rows: List[Dict[str, Any]],
) -> Dict[str, Any]:
    """
    Aggregate per-team regular-season and playoff summaries from weekly rows.
    """
    from collections import defaultdict

    # Bucket by regular season vs playoffs.
    buckets = {
        "regular": {},
        "playoffs": {},
    }

    for row in weekly_rows:
        is_playoffs = bool(row.get("is_playoffs"))
        bucket_name = "playoffs" if is_playoffs else "regular"
        team_key = str(row.get("team_key") or "")
        if not team_key:
            continue

        b = buckets[bucket_name].setdefault(
            team_key,
            {
                "weeks_played": 0,
                "totals": defaultdict(float),
                "h2h_record": {"wins": 0, "losses": 0, "ties": 0},
            },
        )

        b["weeks_played"] += 1

        # Sum categories
        cats = row.get("categories") or {}
        for stat_id, value in cats.items():
            try:
                v = float(value)
            except (TypeError, ValueError):
                continue
            b["totals"][str(stat_id)] += v

        # Sum H2H results (category wins/losses/ties per week)
        res = row.get("result") or {}
        for k in ("wins", "losses", "ties"):
            try:
                b["h2h_record"][k] += int(res.get(k) or 0)
            except (TypeError, ValueError):
                continue

    def build_per_team(bucket: Dict[str, Any]) -> List[Dict[str, Any]]:
        per_team: List[Dict[str, Any]] = []
        for team_key, data in bucket.items():
            weeks_played = data["weeks_played"] or 0
            totals = {sid: float(v) for sid, v in data["totals"].items()}
            avg_per_week: Dict[str, float] = {}
            if weeks_played > 0:
                for sid, total in totals.items():
                    avg_per_week[sid] = total / weeks_played

            per_team.append(
                {
                    "team_key": team_key,
                    "weeks_played": weeks_played,
                    "totals": totals,
                    "avg_per_week": avg_per_week,
                    "h2h_record": data["h2h_record"],
                    # Power rank to be implemented later (sum-of-ranks / z-score).
                    "power_rank": None,
                }
            )
        return per_team

    def build_per_stat_ranks(bucket: Dict[str, Any]) -> Dict[str, List[Dict[str, Any]]]:
        # For each stat_id, build a ranking across teams based on totals.
        by_stat: Dict[str, List[Tuple[str, float]]] = {}
        for team_key, data in bucket.items():
            totals = data["totals"]
            for sid, total in totals.items():
                by_stat.setdefault(str(sid), []).append((team_key, float(total)))

        ranks_out: Dict[str, List[Dict[str, Any]]] = {}
        for sid, pairs in by_stat.items():
            # Higher is better for now; we can add per-stat direction later.
            pairs.sort(key=lambda x: x[1], reverse=True)
            ranked: List[Dict[str, Any]] = []
            rank = 1
            for team_key, value in pairs:
                ranked.append(
                    {
                        "team_key": team_key,
                        "rank": rank,
                        "value": value,
                    }
                )
                rank += 1
            ranks_out[sid] = ranked

        return ranks_out

    regular_bucket = buckets["regular"]
    playoff_bucket = buckets["playoffs"]

    regular_section = {
        "per_team": build_per_team(regular_bucket),
        "per_stat_ranks": build_per_stat_ranks(regular_bucket),
    }
    playoffs_section = {
        "per_team": build_per_team(playoff_bucket),
        "per_stat_ranks": build_per_stat_ranks(playoff_bucket),
        # Bracket / rounds reconstruction can be added later.
        "rounds": [],
        "final_standings": [],
    }

    return {
        "league_key": league_key,
        "regular_season": regular_section,
        "playoffs": playoffs_section,
    }


# ---------------- Excel writer ----------------


def _to_excel(
    league_ctx: LeagueContext,
    weekly_rows: List[Dict[str, Any]],
    matchup_summaries: List[Dict[str, Any]],
    xlsx_path: Path,
    run_ts: Optional[RunTimestamps] = None,
) -> None:
    """
    Write an Excel workbook summarizing matchups, weekly totals, and summaries.
    """
    from collections import defaultdict

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # 1) Matchups sheet
    ws_matchups = wb.active
    ws_matchups.title = "Matchups"
    headers_m = [
        "week",
        "is_playoffs",
        "is_consolation",
        "team_a_key",
        "team_a_name",
        "team_a_points",
        "team_b_key",
        "team_b_name",
        "team_b_points",
        "winner_team_key",
        "winner_team_name",
        "is_tied",
    ]
    ws_matchups.append(headers_m)

    team_name_by_key = {str(t.get("team_key")): t.get("name") for t in league_ctx.teams}

    for m in sorted(matchup_summaries, key=lambda x: (x.get("week"), x.get("team_a_key", ""))):
        ta_key = m.get("team_a_key")
        tb_key = m.get("team_b_key")
        winner_key = m.get("winner_team_key")
        ws_matchups.append(
            [
                m.get("week"),
                bool(m.get("is_playoffs")),
                bool(m.get("is_consolation")),
                ta_key,
                team_name_by_key.get(str(ta_key), ta_key),
                m.get("team_a_points"),
                tb_key,
                team_name_by_key.get(str(tb_key), tb_key),
                m.get("team_b_points"),
                winner_key,
                team_name_by_key.get(str(winner_key), winner_key),
                bool(m.get("is_tied")),
            ]
        )

    ws_matchups.freeze_panes = "A2"
    ws_matchups.auto_filter.ref = f"A1:{get_column_letter(ws_matchups.max_column)}{ws_matchups.max_row}"
    for col_idx in range(1, ws_matchups.max_column + 1):
        ws_matchups.column_dimensions[get_column_letter(col_idx)].width = 16

    # 2) Weekly totals sheet
    ws_weekly = wb.create_sheet("WeeklyTotals")

    # Determine stat IDs present and order them based on league settings.
    scoring = league_ctx.scoring or {}
    stat_categories = scoring.get("stat_categories") or []
    stat_id_order: List[str] = []
    for cat in stat_categories:
        sid = cat.get("stat_id")
        if sid is None:
            continue
        stat_id_order.append(str(sid))

    # Some leagues might have additional stats present in weekly rows only.
    present_ids = set()
    for row in weekly_rows:
        for sid in (row.get("categories") or {}).keys():
            present_ids.add(str(sid))
    for sid in sorted(present_ids):
        if sid not in stat_id_order:
            stat_id_order.append(sid)

    # Build label mapping: stat_id -> display_name (fallback to name).
    stat_labels: Dict[str, str] = {}
    for cat in stat_categories:
        sid = str(cat.get("stat_id"))
        label = cat.get("display_name") or cat.get("abbr") or cat.get("name") or sid
        stat_labels[sid] = str(label)

    headers_w = [
        "week",
        "is_playoffs",
        "is_consolation",
        "team_key",
        "team_name",
        "opponent_key",
        "opponent_name",
        "prev_opponent_key",
    ]
    for sid in stat_id_order:
        headers_w.append(f"stat_{sid}_{stat_labels.get(sid, sid)}")
    headers_w.extend(["cats_wins", "cats_losses", "cats_ties", "team_points"])

    ws_weekly.append(headers_w)

    for row in sorted(weekly_rows, key=lambda x: (x.get("week"), x.get("team_key", ""))):
        team_key = str(row.get("team_key") or "")
        opp_key = str(row.get("opponent_key") or "")
        prev_opp = row.get("prev_opponent_key")
        cats = row.get("categories") or {}
        res = row.get("result") or {}
        out_row = [
            row.get("week"),
            bool(row.get("is_playoffs")),
            bool(row.get("is_consolation")),
            team_key,
            team_name_by_key.get(team_key, team_key),
            opp_key,
            team_name_by_key.get(opp_key, opp_key),
            prev_opp,
        ]
        for sid in stat_id_order:
            val = cats.get(sid)
            out_row.append(val)
        out_row.extend(
            [
                res.get("wins"),
                res.get("losses"),
                res.get("ties"),
                row.get("team_points"),
            ]
        )
        ws_weekly.append(out_row)

    ws_weekly.freeze_panes = "A2"
    ws_weekly.auto_filter.ref = f"A1:{get_column_letter(ws_weekly.max_column)}{ws_weekly.max_row}"
    for col_idx in range(1, 9):
        ws_weekly.column_dimensions[get_column_letter(col_idx)].width = 18
    # Stat columns
    for col_idx in range(9, ws_weekly.max_column + 1):
        ws_weekly.column_dimensions[get_column_letter(col_idx)].width = 12

    # 3) Summary sheets (regular + playoffs) – optional, but useful.
    summary = _aggregate_summary(league_ctx.league_key, weekly_rows)

    def write_summary_sheet(name: str, bucket_key: str) -> None:
        section = summary.get(bucket_key) or {}
        per_team = section.get("per_team") or []
        per_stat_ranks = section.get("per_stat_ranks") or {}

        ws = wb.create_sheet(name)
        # Header: team-level aggregates
        ws.append(["team_key", "team_name", "weeks_played", "wins", "losses", "ties", "power_rank"])
        for row in per_team:
            team_key = row.get("team_key")
            rec = row.get("h2h_record") or {}
            ws.append(
                [
                    team_key,
                    team_name_by_key.get(str(team_key), team_key),
                    row.get("weeks_played"),
                    rec.get("wins"),
                    rec.get("losses"),
                    rec.get("ties"),
                    row.get("power_rank"),
                ]
            )
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 16

        # Optional: second table with per-stat ranks, appended below a blank row.
        if per_stat_ranks:
            ws.append([])
            ws.append(["stat_id", "rank", "team_key", "team_name", "value"])
            for sid, entries in per_stat_ranks.items():
                for entry in entries:
                    tkey = entry.get("team_key")
                    ws.append(
                        [
                            sid,
                            entry.get("rank"),
                            tkey,
                            team_name_by_key.get(str(tkey), tkey),
                            entry.get("value"),
                        ]
                    )

    write_summary_sheet("RegularSummary", "regular_season")
    write_summary_sheet("PlayoffSummary", "playoffs")

    # 4) Run info sheet
    if run_ts is not None:
        ws_run = wb.create_sheet("RunInfo")
        ws_run.append(["field", "value"])
        ws_run.append(["_generated_iso_utc", run_ts.iso_utc])
        ws_run.append(["_generated_iso_local", run_ts.iso_local])
        ws_run.append(["_generated_excel_serial", run_ts.excel_serial])

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


# ---------------- CLI + main ----------------


def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Dump Yahoo Fantasy week-by-week standings (matchups + per-team weekly totals)."
    )
    g = p.add_mutually_exclusive_group(required=True)
    g.add_argument("--league-key", help="Full league key, e.g., nhl.453.l.33099")
    g.add_argument("--league-id", type=int, help="League ID (use with --game)")

    p.add_argument("--game", default="nhl", help="Game key (default: nhl)")
    p.add_argument("--pretty", action="store_true", help="Pretty-print JSON outputs")
    p.add_argument("--to-excel", action="store_true", help="Also write an Excel workbook")

    p.add_argument("--since-week", type=int, help="First week to include (defaults to league start_week)")
    p.add_argument("--through-week", type=int, help="Last week to include (defaults to league end_week)")

    g2 = p.add_mutually_exclusive_group()
    g2.add_argument(
        "--include-playoffs",
        dest="include_playoffs",
        action="store_true",
        help="Include playoff weeks (default)",
    )
    g2.add_argument(
        "--regular-season-only",
        dest="regular_season_only",
        action="store_true",
        help="Exclude playoff weeks (regular season only)",
    )
    p.set_defaults(include_playoffs=True)

    # For now opponent columns are always included; this flag exists for future use.
    p.add_argument(
        "--include-opponent-cols",
        action="store_true",
        default=True,
        help="Include opponent/prev_opponent columns in weekly JSON (default: on)",
    )

    return p.parse_args()


def _resolve_league_key(args: argparse.Namespace) -> str:
    if args.league_key:
        return args.league_key
    return f"{args.game}.l.{args.league_id}"


def main() -> None:
    args = _parse_args()
    league_key = _resolve_league_key(args)

    # Load league context from latest league_dump processed JSON.
    league_ctx = _load_latest_league_dump(league_key)

    # Determine week range from league info + overrides.
    league_info = league_ctx.league_info
    try:
        start_week = int(league_info.get("start_week"))
        end_week = int(league_info.get("end_week"))
    except (TypeError, ValueError):
        print("[ERROR] league_info.start_week / end_week missing or invalid in league_dump output.")
        raise SystemExit(1)

    since_week = args.since_week if args.since_week is not None else start_week
    through_week = args.through_week if args.through_week is not None else end_week

    # Clamp to league range.
    if since_week < start_week:
        since_week = start_week
    if through_week > end_week:
        through_week = end_week

    if since_week > through_week:
        print(
            f"[ERROR] Invalid week range: since-week={since_week}, through-week={through_week} "
            f"(league weeks: {start_week}-{end_week})"
        )
        raise SystemExit(1)

    # Playoff inclusion/exclusion
    include_playoffs = args.include_playoffs and not args.regular_season_only
    playoff_start_week_str = (league_ctx.scoring.get("head_to_head") or {}).get("playoff_start_week")
    try:
        playoff_start_week = int(playoff_start_week_str) if playoff_start_week_str is not None else None
    except (TypeError, ValueError):
        playoff_start_week = None

    # Prepare directory layout for this league.
    run_ts = make_run_timestamps()
    paths = _prepare_standings_dirs(league_key)
    iso_stamp = run_ts.iso_stamp

    print(f"Using standings_dump layout under {paths.root}")

    # Determine which stat_ids are scored (exclude "only_display" stats).
    scoring = league_ctx.scoring or {}
    stat_categories = scoring.get("stat_categories") or []
    scoring_stat_ids: List[str] = []
    for cat in stat_categories:
        sid = cat.get("stat_id")
        if sid is None:
            continue
        is_only_display = str(cat.get("is_only_display_stat") or "").strip()
        if is_only_display in ("1", "true", "True"):
            # Skip display-only stats (e.g., raw shots against where GAA is used).
            continue
        scoring_stat_ids.append(str(sid))

    raw_paths: List[Path] = []
    matchup_summaries: List[Dict[str, Any]] = []
    weekly_rows: List[Dict[str, Any]] = []

    # Iterate weeks and pull scoreboard data.
    for week in range(since_week, through_week + 1):
        try:
            payload = _fetch_scoreboard_week(league_key, week)
        except Exception as e:
            print(f"[WARN] Failed to fetch scoreboard for week {week}: {e}")
            continue

        scoreboard_node = _get_scoreboard_node(payload)
        if not scoreboard_node:
            print(f"[INFO] No scoreboard available for week {week}; skipping.")
            continue

        # Raw snapshot for this week.
        raw_path = paths.raw_dir / f"scoreboard.wk{week:03d}.{iso_stamp}.json"
        with raw_path.open("w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2 if args.pretty else None)
        raw_paths.append(raw_path)
        print(f"Wrote raw scoreboard for week {week}: {raw_path}")

        # Parse all matchups in this week.
        week_matchups: List[Dict[str, Any]] = []
        week_weekly_rows: List[Dict[str, Any]] = []
        is_week_playoffs = False

        for matchup in _iter_matchups(scoreboard_node):
            summary, rows = _compute_weekly_results_for_matchup(matchup, scoring_stat_ids)
            if not summary:
                continue
            week_matchups.append(summary)
            week_weekly_rows.extend(rows)
            if summary.get("is_playoffs"):
                is_week_playoffs = True

        # Apply playoff filters.
        if is_week_playoffs and not include_playoffs:
            print(f"[INFO] Week {week} is playoffs and --regular-season-only is active; skipping week.")
            continue

        if not week_matchups:
            print(f"[INFO] No usable matchups parsed for week {week}; skipping.")
            continue

        # Collect for this run.
        matchup_summaries.extend(week_matchups)
        weekly_rows.extend(week_weekly_rows)

    if not weekly_rows:
        print("[ERROR] No weekly rows were produced; nothing to write.")
        raise SystemExit(1)

    # Backfill prev_opponent_key now that all weeks are loaded.
    _backfill_prev_opponents(weekly_rows)

    # Build processed JSON payloads.
    matchups_payload = {
        "league_key": league_key,
        "_generated_unix": run_ts.unix,
        "_generated_iso_utc": run_ts.iso_utc,
        "_generated_iso_local": run_ts.iso_local,
        "weeks": [],
    }

    # Group matchup summaries by week.
    from collections import defaultdict

    by_week: Dict[int, Dict[str, Any]] = defaultdict(lambda: {"matchups": []})
    for ms in matchup_summaries:
        w = int(ms.get("week"))
        bw = by_week[w]
        bw.setdefault("week", w)
        bw.setdefault("week_start", ms.get("week_start"))
        bw.setdefault("week_end", ms.get("week_end"))
        # If any matchup is playoffs, treat week as playoffs.
        if ms.get("is_playoffs"):
            bw["is_playoffs"] = True
        else:
            bw.setdefault("is_playoffs", False)
        bw["matchups"].append(ms)

    for w in sorted(by_week.keys()):
        matchups_payload["weeks"].append(by_week[w])

    weekly_payload = {
        "league_key": league_key,
        "_generated_unix": run_ts.unix,
        "_generated_iso_utc": run_ts.iso_utc,
        "_generated_iso_local": run_ts.iso_local,
        "rows": weekly_rows,
    }

    summary_payload = _aggregate_summary(league_key, weekly_rows)
    summary_payload["_generated_unix"] = run_ts.unix
    summary_payload["_generated_iso_utc"] = run_ts.iso_utc
    summary_payload["_generated_iso_local"] = run_ts.iso_local

    # Write processed JSON files.
    matchups_path = paths.processed_dir / f"matchups.{iso_stamp}.json"
    weekly_path = paths.processed_dir / f"weekly.{iso_stamp}.json"
    summary_path = paths.processed_dir / f"summary.{iso_stamp}.json"

    for path, payload in (
        (matchups_path, matchups_payload),
        (weekly_path, weekly_payload),
        (summary_path, summary_payload),
    ):
        with path.open("w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2 if args.pretty else None)
        print(f"Wrote processed: {path}")

    produced_paths: List[Path] = []
    produced_paths.extend(raw_paths)
    produced_paths.extend([matchups_path, weekly_path, summary_path])

    # Optional Excel workbook
    if args.to_excel:
        week_range_label = f"wk{since_week:02d}-{through_week:02d}"
        excel_path = paths.excel_dir / f"standings.{week_range_label}.{iso_stamp}.xlsx"
        _to_excel(league_ctx, weekly_rows, matchup_summaries, excel_path, run_ts=run_ts)
        print(f"Wrote Excel: {excel_path}")
        produced_paths.append(excel_path)
        excel_rel: Optional[str] = excel_path.relative_to(paths.root).as_posix()
    else:
        excel_rel = None

    # Update _meta/latest.json for standings_dump.
    matchups_rel = matchups_path.relative_to(paths.root).as_posix()
    weekly_rel = weekly_path.relative_to(paths.root).as_posix()
    summary_rel = summary_path.relative_to(paths.root).as_posix()
    latest_path = _update_latest(paths, run_ts, matchups_rel, weekly_rel, summary_rel, excel_rel)
    print(f"Updated latest.json: {latest_path}")
    produced_paths.append(latest_path)

    # Manifest for this run.
    cli_args: Dict[str, Any] = {
        "league_key": league_key,
        "league_id": args.league_id,
        "game": args.game,
        "pretty": args.pretty,
        "to_excel": args.to_excel,
        "since_week": since_week,
        "through_week": through_week,
        "include_playoffs": include_playoffs,
    }
    manifest_data = _build_manifest_dict(
        module_name="standings_dump",
        league_key=league_key,
        paths=paths,
        run_ts=run_ts,
        cli_args=cli_args,
        produced_paths=produced_paths,
    )
    manifest_path = _write_manifest(paths, run_ts, manifest_data)
    print(f"Wrote manifest: {manifest_path}")


if __name__ == "__main__":
    main()
