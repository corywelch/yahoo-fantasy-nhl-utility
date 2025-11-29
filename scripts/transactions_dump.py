#!/usr/bin/env python3
from __future__ import annotations

"""Transactions dump: full-season move ledger (adds/drops/trades/etc).

This script depends on a prior league_dump run for the same league, and will
refuse to run if it cannot locate the latest processed league JSON under:

  exports/<league_key>/_meta/latest.json   (league_dump.processed)

Outputs under:

  exports/<league_key>/transactions_dump/
    raw/transactions.<ISO>.json
    processed/master.<ISO>.json
    excel/transactions.<ISO>.xlsx  (when --to-excel)
    manifest/manifest.<ISO>.json

Usage example:

    python -m scripts.transactions_dump --league-key 465.l.22607 --to-excel --pretty
"""

import argparse
import json
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import requests

from src.auth.oauth import get_session
from src.config.env import get_export_dir
from src.util_time import RunTimestamps, make_run_timestamps

BASE_URL = "https://fantasysports.yahooapis.com/fantasy/v2"


@dataclass
class WeekRange:
    week: int
    start_date: str  # YYYY-MM-DD
    end_date: str    # YYYY-MM-DD
    is_playoffs: Optional[bool] = None


@dataclass
class Paths:
    league_root: Path
    meta_dir: Path
    raw_dir: Path
    processed_dir: Path
    excel_dir: Path
    manifest_dir: Path


def _paths_for_league(league_key: str) -> Paths:
    root = get_export_dir() / league_key
    td_root = root / "transactions_dump"
    return Paths(
        league_root=root,
        meta_dir=root / "_meta",
        raw_dir=td_root / "raw",
        processed_dir=td_root / "processed",
        excel_dir=td_root / "excel",
        manifest_dir=td_root / "manifest",
    )


def _ensure_dirs(paths: Paths) -> None:
    for d in (paths.meta_dir, paths.raw_dir, paths.processed_dir, paths.excel_dir, paths.manifest_dir):
        d.mkdir(parents=True, exist_ok=True)


def _datetime_to_excel_serial(dt: datetime) -> float:
    """Convert a UTC datetime to an Excel serial (1900 date system).

    Excel's epoch is 1899-12-30 for modern versions; dates after 1900-03-01
    can be represented accurately with this simple offset.
    """
    excel_epoch = datetime(1899, 12, 30, tzinfo=timezone.utc)
    delta = dt - excel_epoch
    return delta.days + (delta.seconds + delta.microseconds / 1_000_000.0) / 86400.0


def _fetch_json(session: requests.Session, path: str) -> Dict[str, Any]:
    url = f"{BASE_URL}/{path}?format=json"
    resp = session.get(url)
    resp.raise_for_status()
    return resp.json()


def _load_league_context(paths: Paths) -> Tuple[Dict[str, Any], str]:
    """Load the latest processed league_dump JSON via _meta/latest.json.

    Returns (league_dump_json, processed_rel_path).

    Exits with code 1 and a clear message if the prerequisite league_dump
    output cannot be found.
    """
    latest_path = paths.meta_dir / "latest.json"
    if not latest_path.exists():
        print(
            "ERROR: Expected exports/%s/_meta/latest.json but it is missing.\n"
            "Run league_dump for this league before running transactions_dump."
            % paths.league_root.name,
            file=sys.stderr,
        )
        sys.exit(1)

    try:
        with latest_path.open("r", encoding="utf-8") as f:
            latest = json.load(f)
    except Exception as exc:  # pragma: no cover - defensive
        print(
            f"ERROR: Failed to parse {latest_path} ({exc}).\n"
            "Re-run league_dump to regenerate metadata.",
            file=sys.stderr,
        )
        sys.exit(1)

    league_block = latest.get("league_dump")
    if not isinstance(league_block, dict) or "processed" not in league_block:
        print(
            "ERROR: _meta/latest.json does not contain 'league_dump.processed'.\n"
            "Run league_dump for this league before running transactions_dump.",
            file=sys.stderr,
        )
        sys.exit(1)

    processed_rel = league_block["processed"]
    processed_path = paths.league_root / processed_rel
    if not processed_path.exists():
        print(
            f"ERROR: Processed league_dump JSON not found at '{processed_rel}'.\n"
            "Re-run league_dump for this league and try again.",
            file=sys.stderr,
        )
        sys.exit(1)

    try:
        with processed_path.open("r", encoding="utf-8") as f:
            league_dump = json.load(f)
    except Exception as exc:  # pragma: no cover - defensive
        print(
            f"ERROR: Failed to parse processed league JSON at {processed_rel} ({exc}).",
            file=sys.stderr,
        )
        sys.exit(1)

    if "league_info" not in league_dump or "scoring" not in league_dump or "teams" not in league_dump:
        print(
            "ERROR: Processed league_dump JSON is missing expected keys "
            "('league_info', 'scoring', 'teams').",
            file=sys.stderr,
        )
        sys.exit(1)

    return league_dump, processed_rel


def _build_week_index(
    session: requests.Session,
    league_key: str,
    start_week: int,
    end_week: int,
    head_to_head: Dict[str, Any],
) -> List[WeekRange]:
    """Build a list of WeekRange from scoreboard;week=N for each week.

    We only use this to map transaction timestamps → matchup weeks. If any
    particular week fails to fetch, it is skipped; transactions that cannot
    be mapped will have week=None.
    """
    ranges: List[WeekRange] = []

    for wk in range(start_week, end_week + 1):
        path = f"league/{league_key}/scoreboard;week={wk}"
        try:
            payload = _fetch_json(session, path)
        except requests.HTTPError as exc:
            print(f"WARNING: Failed to fetch scoreboard for week {wk}: {exc}", file=sys.stderr)
            continue

        fc = payload.get("fantasy_content", {})
        league = fc.get("league")
        if not isinstance(league, list) or len(league) < 2:
            continue
        sb = league[1].get("scoreboard")
        if not isinstance(sb, dict):
            continue
        sb_week = sb.get("week")
        root = sb.get("0")
        if not isinstance(root, dict):
            continue
        matchups = root.get("matchups")
        if not isinstance(matchups, dict):
            continue

        meta = None
        for mk, mv in matchups.items():
            if mk == "count":
                continue
            matchup = mv.get("matchup")
            if isinstance(matchup, list) and matchup:
                meta = matchup[0]
            elif isinstance(matchup, dict):
                meta = matchup
            if isinstance(meta, dict):
                break

        if not isinstance(meta, dict):
            continue

        start_date = meta.get("week_start")
        end_date = meta.get("week_end")
        if not start_date or not end_date:
            continue

        is_playoffs_raw = meta.get("is_playoffs")
        is_playoffs: Optional[bool]
        if isinstance(is_playoffs_raw, str):
            is_playoffs = is_playoffs_raw == "1"
        elif isinstance(is_playoffs_raw, (int, bool)):
            is_playoffs = bool(is_playoffs_raw)
        else:
            is_playoffs = None

        try:
            week_num = int(sb_week) if sb_week is not None else wk
        except Exception:
            week_num = wk

        ranges.append(WeekRange(week=week_num, start_date=start_date, end_date=end_date, is_playoffs=is_playoffs))

    # Normalize playoff flags using head_to_head if necessary
    h2h_playoff_start = None
    h2h = head_to_head or {}
    if isinstance(h2h, dict) and h2h.get("uses_playoff") == "1" and "playoff_start_week" in h2h:
        try:
            h2h_playoff_start = int(h2h["playoff_start_week"])
        except Exception:
            h2h_playoff_start = None

    if h2h_playoff_start is not None:
        for wr in ranges:
            if wr.week >= h2h_playoff_start:
                wr.is_playoffs = True if wr.is_playoffs is None else wr.is_playoffs
            else:
                if wr.is_playoffs is None:
                    wr.is_playoffs = False

    ranges.sort(key=lambda wr: wr.week)
    return ranges


def _lookup_week(week_index: Iterable[WeekRange], date_str: str) -> Tuple[Optional[int], Optional[bool]]:
    """Return (week, is_playoffs) for a YYYY-MM-DD date string, if known."""
    for wr in week_index:
        if wr.start_date <= date_str <= wr.end_date:
            return wr.week, wr.is_playoffs
    return None, None


def _normalize_transactions(
    raw_payload: Dict[str, Any],
    league_dump: Dict[str, Any],
    week_index: List[WeekRange],
    run_ts: RunTimestamps,
    type_filter: Optional[Iterable[str]],
    include_meta: bool,
) -> Dict[str, Any]:
    """Normalize Yahoo transactions payload into a tidy JSON structure."""
    fc = raw_payload.get("fantasy_content", {})
    league = fc.get("league")
    if not isinstance(league, list) or len(league) < 2:
        raise ValueError("Unexpected transactions payload shape: missing league list")

    league_info = league_dump["league_info"]
    scoring = league_dump["scoring"]
    teams = league_dump["teams"]

    league_key = league_info.get("league_key")
    season = league_info.get("season")

    tx_container = league[1].get("transactions", {})
    if not isinstance(tx_container, dict):
        raise ValueError("Unexpected transactions payload shape: league[1].transactions missing")

    # Normalize type filter (header-level types)
    norm_filter: Optional[set] = None
    if type_filter is not None:
        norm_filter = set()
        for t in type_filter:
            t = (t or "").strip()
            if not t:
                continue
            if t.lower() in {"add/drop", "add_drop"}:
                norm_filter.add("add_drop")
            else:
                norm_filter.add(t)

    def _extract_team_name(team_key: Optional[str]) -> Optional[str]:
        if not team_key:
            return None
        for t in teams:
            if t.get("team_key") == team_key:
                return t.get("name")
        return None

    week_index_list = list(week_index)

    transactions_out: List[Dict[str, Any]] = []

    for k, v in tx_container.items():
        if k == "count":
            continue
        tx = v.get("transaction")
        if not isinstance(tx, list) or not tx:
            continue

        header: Dict[str, Any] = {}
        players_block: Optional[Dict[str, Any]] = None
        extra_blocks: List[Dict[str, Any]] = []

        for part in tx:
            if not isinstance(part, dict):
                continue
            if "players" in part:
                players_block = part.get("players")
            elif "timestamp" in part and "transaction_id" in part:
                header = part
            else:
                extra_blocks.append(part)

        if not header and isinstance(tx[0], dict):
            header = tx[0]

        if not header:
            continue  # nothing we can do

        raw_type = header.get("type")
        if raw_type is None:
            continue
        norm_type = "add_drop" if raw_type == "add/drop" else raw_type

        if norm_filter is not None and norm_type not in norm_filter:
            continue

        # Timestamp + week mapping
        ts_unix: Optional[int]
        ts_iso_utc: Optional[str]
        ts_excel: Optional[float]
        week_val: Optional[int]
        is_playoffs: Optional[bool]

        ts_str = header.get("timestamp")
        if ts_str is not None:
            try:
                ts_unix = int(ts_str)
            except Exception:
                ts_unix = None
        else:
            ts_unix = None

        if ts_unix is not None:
            dt_utc = datetime.utcfromtimestamp(ts_unix).replace(tzinfo=timezone.utc)
            ts_iso_utc = dt_utc.isoformat().replace("+00:00", "Z")
            ts_excel = _datetime_to_excel_serial(dt_utc)
            date_str = dt_utc.date().isoformat()
            week_val, is_playoffs = _lookup_week(week_index_list, date_str)
        else:
            ts_iso_utc = None
            ts_excel = None
            week_val, is_playoffs = None, None

        # Moves
        moves: List[Dict[str, Any]] = []

        if isinstance(players_block, dict):
            for pidx, pnode in players_block.items():
                if pidx == "count":
                    continue
                if not isinstance(pnode, dict):
                    continue
                pdata = pnode.get("player")
                if not isinstance(pdata, list) or not pdata:
                    continue

                base_attrs = pdata[0]
                if not isinstance(base_attrs, list):
                    base_attrs = [base_attrs]

                player_key = None
                player_id = None
                name = None
                editorial_team_abbr = None
                display_position = None

                for item in base_attrs:
                    if not isinstance(item, dict):
                        continue
                    if "player_key" in item:
                        player_key = item.get("player_key")
                    elif "player_id" in item:
                        player_id = item.get("player_id")
                    elif "name" in item:
                        name = item.get("name")
                    elif "editorial_team_abbr" in item:
                        editorial_team_abbr = item.get("editorial_team_abbr")
                    elif "display_position" in item:
                        display_position = item.get("display_position")

                # transaction_data may be a dict or list of dicts
                td_block = None
                for extra in pdata[1:]:
                    if isinstance(extra, dict) and "transaction_data" in extra:
                        td_block = extra.get("transaction_data")
                        break

                if td_block is None:
                    continue

                if isinstance(td_block, dict):
                    td_list = [td_block]
                elif isinstance(td_block, list):
                    td_list = [td for td in td_block if isinstance(td, dict)]
                else:
                    td_list = []

                for td in td_list:
                    move_type = td.get("type")
                    source_type = td.get("source_type")
                    dest_type = td.get("destination_type")

                    from_team_key = td.get("source_team_key") if source_type == "team" else None
                    to_team_key = td.get("destination_team_key") if dest_type == "team" else None

                    from_team_name = td.get("source_team_name") or _extract_team_name(from_team_key)
                    to_team_name = td.get("destination_team_name") or _extract_team_name(to_team_key)

                    via: Optional[str] = None
                    if move_type == "add":
                        if source_type == "freeagents":
                            via = "free_agent"
                        elif source_type == "waivers":
                            via = "waivers"
                        elif source_type == "team":
                            via = "trade"
                    elif move_type == "drop":
                        if dest_type == "waivers":
                            via = "waivers"
                        elif dest_type == "team":
                            via = "trade"
                    elif move_type == "trade":
                        via = "trade"

                    move = {
                        "player_key": player_key,
                        "player_id": player_id,
                        "player_name": name.get("full") if isinstance(name, dict) else None,
                        "editorial_team_abbr": editorial_team_abbr,
                        "display_position": display_position,
                        "transaction_player_type": move_type,
                        "from_team_key": from_team_key,
                        "from_team_name": from_team_name,
                        "to_team_key": to_team_key,
                        "to_team_name": to_team_name,
                        "source_type": source_type,
                        "destination_type": dest_type,
                        "via": via,
                        "waiver_priority_before": td.get("waiver_priority_before"),
                        "waiver_priority_after": td.get("waiver_priority_after"),
                        "faab_bid": td.get("faab_bid"),
                        "faab_winning": td.get("faab_winning"),
                    }
                    moves.append(move)

        norm_tx: Dict[str, Any] = {
            "transaction_id": header.get("transaction_id"),
            "transaction_key": header.get("transaction_key"),
            "type": norm_type,
            "status": header.get("status"),
            "timestamp_unix": ts_unix,
            "timestamp_iso_utc": ts_iso_utc,
            "timestamp_excel": ts_excel,
            "week": week_val,
            "is_playoffs": is_playoffs,
            "initiating_team_key": None,  # not exposed reliably by Yahoo
            "manager_move_count_week": None,  # reserved for future
            "moves_remaining_estimate": None,  # reserved for future
            "notes": None,
            "moves": moves,
            "fetched_unix": run_ts.unix,
            "fetched_iso_utc": run_ts.iso_utc,
            "fetched_excel_serial": run_ts.excel_serial,
        }

        if include_meta:
            meta_block: Dict[str, Any] = {"header": header}
            # if there is a picks block (trades), keep it
            for blk in extra_blocks:
                if "picks" in blk:
                    meta_block["picks"] = blk.get("picks")
            norm_tx["meta"] = meta_block

        transactions_out.append(norm_tx)

    # Sort by timestamp then transaction_id for stable output
    transactions_out.sort(key=lambda tx: (tx.get("timestamp_unix") or 0, int(tx.get("transaction_id") or 0)))

    out: Dict[str, Any] = {
        "league_key": league_key,
        "season": season,
        "fetched_unix": run_ts.unix,
        "fetched_iso_utc": run_ts.iso_utc,
        "fetched_iso_local": run_ts.iso_local,
        "week_index": [wr.__dict__ for wr in week_index_list],
        "transaction_types_included": sorted(norm_filter) if norm_filter is not None else "all",
        "transactions": transactions_out,
    }
    return out


def _dump_json(data: Any, path: Path, pretty: bool) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        if pretty:
            json.dump(data, f, indent=2, sort_keys=False)
        else:
            json.dump(data, f, separators=(",", ":"), sort_keys=False)


def _write_excel(transactions: List[Dict[str, Any]], path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(
            "ERROR: openpyxl is not installed but --to-excel was requested.\n"
            "Install it with 'pip install openpyxl' and try again.",
            file=sys.stderr,
        )
        sys.exit(1)

    path.parent.mkdir(parents=True, exist_ok=True)

    # Flatten moves into rows
    rows_all: List[Dict[str, Any]] = []
    for tx in transactions:
        base = {
            "transaction_id": tx.get("transaction_id"),
            "transaction_key": tx.get("transaction_key"),
            "type": tx.get("type"),
            "status": tx.get("status"),
            "timestamp_iso_utc": tx.get("timestamp_iso_utc"),
            "week": tx.get("week"),
            "is_playoffs": tx.get("is_playoffs"),
        }
        for mv in tx.get("moves", []):
            row = dict(base)
            row.update(
                {
                    "player_key": mv.get("player_key"),
                    "player_name": mv.get("player_name"),
                    "player_id": mv.get("player_id"),
                    "editorial_team_abbr": mv.get("editorial_team_abbr"),
                    "display_position": mv.get("display_position"),
                    "move_type": mv.get("transaction_player_type"),
                    "via": mv.get("via"),
                    "from_team_key": mv.get("from_team_key"),
                    "from_team_name": mv.get("from_team_name"),
                    "to_team_key": mv.get("to_team_key"),
                    "to_team_name": mv.get("to_team_name"),
                    "waiver_priority_before": mv.get("waiver_priority_before"),
                    "waiver_priority_after": mv.get("waiver_priority_after"),
                    "faab_bid": mv.get("faab_bid"),
                    "faab_winning": mv.get("faab_winning"),
                }
            )
            rows_all.append(row)

    # Define a common header ordering
    headers = [
        "transaction_id",
        "transaction_key",
        "type",
        "status",
        "timestamp_iso_utc",
        "week",
        "is_playoffs",
        "player_key",
        "player_name",
        "player_id",
        "editorial_team_abbr",
        "display_position",
        "move_type",
        "via",
        "from_team_key",
        "from_team_name",
        "to_team_key",
        "to_team_name",
        "waiver_priority_before",
        "waiver_priority_after",
        "faab_bid",
        "faab_winning",
    ]

    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "AllMoves"

    def write_sheet(sheet_name: str, rows: List[Dict[str, Any]]) -> None:
        ws = wb.create_sheet(title=sheet_name) if sheet_name != "AllMoves" else ws_all
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h) for h in headers])
        ws.freeze_panes = "A2"
        # Auto-size columns a bit
        for idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].width = 18

    # All moves
    write_sheet("AllMoves", rows_all)

    # Adds
    rows_adds = [r for r in rows_all if r.get("move_type") == "add"]
    if rows_adds:
        write_sheet("Adds", rows_adds)

    # Drops
    rows_drops = [r for r in rows_all if r.get("move_type") == "drop"]
    if rows_drops:
        write_sheet("Drops", rows_drops)

    # Trades
    rows_trades = [r for r in rows_all if r.get("move_type") == "trade"]
    if rows_trades:
        write_sheet("Trades", rows_trades)

    # Remove default empty sheet if we accidentally created one
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        std = wb["Sheet"]
        wb.remove(std)

    wb.save(path)


def _sha256_of_file(path: Path) -> str:
    import hashlib

    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _write_manifest(
    paths: Paths,
    league_key: str,
    run_ts: RunTimestamps,
    cli_args: Dict[str, Any],
    raw_path: Path,
    processed_path: Path,
    excel_path: Optional[Path],
) -> Path:
    paths.manifest_dir.mkdir(parents=True, exist_ok=True)
    manifest_path = paths.manifest_dir / f"manifest.{run_ts.iso_stamp}.json"

    def rel(p: Path) -> str:
        return p.relative_to(paths.league_root).as_posix()

    files: Dict[str, Dict[str, Any]] = {}
    for p in (raw_path, processed_path):
        if p.exists():
            files[rel(p)] = {
                "size_bytes": p.stat().st_size,
                "sha256": _sha256_of_file(p),
            }
    if excel_path is not None and excel_path.exists():
        files[rel(excel_path)] = {
            "size_bytes": excel_path.stat().st_size,
            "sha256": _sha256_of_file(excel_path),
        }

    manifest = {
        "module": "transactions_dump",
        "league_key": league_key,
        "_generated_unix": run_ts.unix,
        "_generated_iso_utc": run_ts.iso_utc,
        "_generated_iso_local": run_ts.iso_local,
        "cli_args": cli_args,
        "files": files,
    }

    _dump_json(manifest, manifest_path, pretty=True)
    return manifest_path


def _update_latest_meta(
    paths: Paths,
    league_key: str,
    processed_path: Path,
    excel_path: Optional[Path],
    run_ts: RunTimestamps,
) -> None:
    latest_path = paths.meta_dir / "latest.json"
    if latest_path.exists():
        try:
            with latest_path.open("r", encoding="utf-8") as f:
                latest = json.load(f)
        except Exception:
            latest = {}
    else:
        latest = {}

    latest["league_key"] = league_key

    rel_processed = processed_path.relative_to(paths.league_root).as_posix()
    rel_excel = (
        excel_path.relative_to(paths.league_root).as_posix()
        if excel_path is not None and excel_path.exists()
        else None
    )

    tx_block: Dict[str, Any] = {"processed": rel_processed}
    if rel_excel is not None:
        tx_block["excel"] = rel_excel

    latest["transactions_dump"] = tx_block
    latest["_updated_unix"] = run_ts.unix
    latest["_updated_iso_utc"] = run_ts.iso_utc

    _dump_json(latest, latest_path, pretty=True)


def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Dump Yahoo Fantasy league transactions for a season.")
    group = p.add_mutually_exclusive_group(required=True)
    group.add_argument("--league-key", help="Full league key, e.g. 465.l.22607")
    group.add_argument("--league-id", help="League ID, e.g. 22607 (paired with --game)")

    p.add_argument("--game", default="nhl", help="Game code (default: nhl)")
    p.add_argument(
        "--types",
        help=(
            "Comma-separated list of transaction header types to include "
            "(e.g. 'add,drop,add_drop,trade,commish'). Default is all."
        ),
    )
    p.add_argument(
        "--include-meta",
        action="store_true",
        help="Include raw header/picks fragments under a 'meta' key for each transaction.",
    )
    p.add_argument("--pretty", action="store_true", help="Pretty-print JSON outputs.")
    p.add_argument("--to-excel", action="store_true", help="Also write an Excel workbook.")
    return p.parse_args()


def _resolve_league_key(args: argparse.Namespace) -> str:
    if args.league_key:
        return args.league_key
    return f"{args.game}.l.{args.league_id}"


def main() -> None:
    args = _parse_args()
    league_key = _resolve_league_key(args)

    session = get_session()

    paths = _paths_for_league(league_key)
    _ensure_dirs(paths)

    league_dump, _league_processed_rel = _load_league_context(paths)
    league_info = league_dump["league_info"]
    scoring = league_dump["scoring"]

    start_week = int(league_info.get("start_week", "1"))
    end_week = int(league_info.get("end_week", start_week))

    head_to_head = scoring.get("head_to_head", {})

    run_ts = make_run_timestamps()

    # Build week index from scoreboard;week=N
    week_index = _build_week_index(session, league_key, start_week, end_week, head_to_head)

    # Fetch transactions
    try:
        raw_payload = _fetch_json(session, f"league/{league_key}/transactions")
    except requests.HTTPError as exc:
        print(f"ERROR: Failed to fetch transactions for {league_key}: {exc}", file=sys.stderr)
        sys.exit(1)

    # Save raw snapshot
    raw_path = paths.raw_dir / f"transactions.{run_ts.iso_stamp}.json"
    _dump_json(raw_payload, raw_path, pretty=args.pretty)
    print(f"Wrote raw transactions JSON: {raw_path}")

    # Normalize
    if args.types:
        type_filter = [t.strip() for t in args.types.split(",") if t.strip()]
    else:
        type_filter = None

    processed = _normalize_transactions(
        raw_payload=raw_payload,
        league_dump=league_dump,
        week_index=week_index,
        run_ts=run_ts,
        type_filter=type_filter,
        include_meta=bool(args.include_meta),
    )

    processed_path = paths.processed_dir / f"master.{run_ts.iso_stamp}.json"
    _dump_json(processed, processed_path, pretty=args.pretty)
    print(f"Wrote processed transactions JSON: {processed_path}")

    excel_path: Optional[Path] = None
    if args.to_excel:
        excel_path = paths.excel_dir / f"transactions.{run_ts.iso_stamp}.xlsx"
        _write_excel(processed.get("transactions", []), excel_path)
        print(f"Wrote Excel workbook: {excel_path}")

    # Manifest + meta update
    cli_args = {
        "league_key": league_key,
        "league_id": args.league_id,
        "game": args.game,
        "types": args.types,
        "include_meta": args.include_meta,
        "pretty": args.pretty,
        "to_excel": args.to_excel,
    }

    manifest_path = _write_manifest(
        paths=paths,
        league_key=league_key,
        run_ts=run_ts,
        cli_args=cli_args,
        raw_path=raw_path,
        processed_path=processed_path,
        excel_path=excel_path,
    )
    print(f"Wrote manifest: {manifest_path}")

    _update_latest_meta(
        paths=paths,
        league_key=league_key,
        processed_path=processed_path,
        excel_path=excel_path,
        run_ts=run_ts,
    )
    print(f"Updated latest metadata: {paths.meta_dir / 'latest.json'}")


if __name__ == "__main__":
    main()
