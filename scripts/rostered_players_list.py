#!/usr/bin/env python3
from __future__ import annotations

"""
rostered_players_list: consolidated player list from draft + transactions.

This script depends on prior runs of ``league_dump``, ``draft_dump``, and
``transactions_dump`` for the same league. It *reads* their latest processed
JSON outputs from:

    exports/<league_key>/_meta/latest.json
      → league_dump.processed
      → draft_dump.processed
      → transactions_dump.processed

and builds a merged list of all players who either:

  - were selected in the league draft, or
  - appeared in at least one season transaction (add/drop/trade/etc).

For each player, the output records:

  - player_key
  - player_name (if available from transactions_dump)
  - drafted_team_key / drafted_team_name (if ever drafted)
  - last_move_type   (one of: drafted, add, drop, trade, or other)
  - last_move_team_key / last_move_team_name (team involved in that move)

Outputs under:

  exports/<league_key>/rostered_players_list/
    processed/
      players.<ISO>.json
    excel/
      players.<ISO>.xlsx        (when --to-excel)
    manifest/
      manifest.<ISO>.json

Where:
  - <league_key> is the full Yahoo league key (e.g. "nhl.465.l.22607")
  - <ISO> is a run identifier like "20251129T014755Z" (UTC timestamp)

This module does not call the Yahoo API directly; it only consolidates
existing processed data from draft_dump and transactions_dump.
"""

import argparse
import json
import sys
import hashlib
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional

from src.config.env import get_export_dir
from src.util_time import RunTimestamps, make_run_timestamps


# ---------------- Data models ----------------


@dataclass
class PlayerSummary:
    player_key: str
    player_name: Optional[str] = None

    drafted_team_key: Optional[str] = None
    drafted_team_name: Optional[str] = None

    last_move_type: Optional[str] = None          # drafted/add/drop/trade/other
    last_move_team_key: Optional[str] = None
    last_move_team_name: Optional[str] = None
    last_move_source: Optional[str] = None        # "draft_dump" or "transactions_dump"
    last_move_timestamp_unix: Optional[float] = None


@dataclass
class Paths:
    league_root: Path
    meta_dir: Path
    processed_dir: Path
    excel_dir: Path
    manifest_dir: Path


# ---------------- Path + IO helpers ----------------


def _paths_for_league(league_key: str) -> Paths:
    root = get_export_dir() / league_key
    rp_root = root / "rostered_players_list"
    return Paths(
        league_root=root,
        meta_dir=root / "_meta",
        processed_dir=rp_root / "processed",
        excel_dir=rp_root / "excel",
        manifest_dir=rp_root / "manifest",
    )


def _ensure_dirs(paths: Paths) -> None:
    paths.meta_dir.mkdir(parents=True, exist_ok=True)
    paths.processed_dir.mkdir(parents=True, exist_ok=True)
    paths.excel_dir.mkdir(parents=True, exist_ok=True)
    paths.manifest_dir.mkdir(parents=True, exist_ok=True)


def _sha256_of_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _dump_json(data: Any, path: Path, pretty: bool) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        if pretty:
            json.dump(data, f, indent=2, sort_keys=False)
        else:
            json.dump(data, f, separators=(",", ":"), sort_keys=False)


# ---------------- Meta + input loading ----------------


def _load_latest_meta(league_key: str, meta_dir: Path) -> Dict[str, Any]:
    latest_path = meta_dir / "latest.json"
    if not latest_path.exists():
        print(
            f"ERROR: _meta/latest.json not found for league {league_key}. "
            "Run league_dump, draft_dump, and transactions_dump first.",
            file=sys.stderr,
        )
        sys.exit(1)

    try:
        with latest_path.open("r", encoding="utf-8") as f:
            latest = json.load(f)
    except Exception as exc:  # pragma: no cover - defensive
        print(f"ERROR: Failed to parse _meta/latest.json ({exc}).", file=sys.stderr)
        sys.exit(1)

    return latest


def _require_block(latest: Dict[str, Any], key: str, league_key: str) -> Dict[str, Any]:
    block = latest.get(key)
    if not isinstance(block, dict) or "processed" not in block:
        print(
            f"ERROR: _meta/latest.json is missing '{key}.processed' for league {league_key}.\n"
            f"       Run {key} for this league before running rostered_players_list.",
            file=sys.stderr,
        )
        sys.exit(1)
    return block


def _load_inputs_for_league(paths: Paths, league_key: str) -> Dict[str, Any]:
    latest = _load_latest_meta(league_key, paths.meta_dir)

    league_block = _require_block(latest, "league_dump", league_key)
    draft_block = _require_block(latest, "draft_dump", league_key)
    tx_block = _require_block(latest, "transactions_dump", league_key)

    league_path = paths.league_root / league_block["processed"]
    draft_path = paths.league_root / draft_block["processed"]
    tx_path = paths.league_root / tx_block["processed"]

    try:
        with league_path.open("r", encoding="utf-8") as f:
            league_dump = json.load(f)
    except Exception as exc:
        print(f"ERROR: Failed to parse league_dump processed JSON at '{league_path}' ({exc}).", file=sys.stderr)
        sys.exit(1)

    try:
        with draft_path.open("r", encoding="utf-8") as f:
            draft_dump = json.load(f)
    except Exception as exc:
        print(f"ERROR: Failed to parse draft_dump processed JSON at '{draft_path}' ({exc}).", file=sys.stderr)
        sys.exit(1)

    try:
        with tx_path.open("r", encoding="utf-8") as f:
            tx_dump = json.load(f)
    except Exception as exc:
        print(
            f"ERROR: Failed to parse transactions_dump processed JSON at '{tx_path}' ({exc}).",
            file=sys.stderr,
        )
        sys.exit(1)

    return {
        "latest": latest,
        "league_dump": league_dump,
        "draft_dump": draft_dump,
        "transactions_dump": tx_dump,
    }


# ---------------- Core merge logic ----------------


def _build_team_name_map(league_dump: Dict[str, Any]) -> Dict[str, str]:
    teams = league_dump.get("teams") or []
    out: Dict[str, str] = {}
    for t in teams:
        if not isinstance(t, dict):
            continue
        tk = t.get("team_key")
        name = t.get("name")
        if tk and isinstance(name, str):
            out[tk] = name
    return out


def _normalize_move_type(raw: Optional[str]) -> str:
    s = (raw or "").lower()
    if s in {"add", "waiver_add"} or s.startswith("add"):
        return "add"
    if s in {"drop", "waiver_drop"} or s.startswith("drop"):
        return "drop"
    if s == "trade":
        return "trade"
    if s in {"draft", "drafted"}:
        return "drafted"
    return s or "other"


def _update_last_move(
    player: PlayerSummary,
    ts_unix: Optional[float],
    move_type: str,
    team_key: Optional[str],
    team_name: Optional[str],
    source: str,
) -> None:
    """Update the 'last move' fields if this event is newer."""
    if ts_unix is None:
        # Only override if we don't have any timestamp yet.
        if player.last_move_timestamp_unix is None:
            player.last_move_timestamp_unix = None
            player.last_move_type = move_type
            player.last_move_team_key = team_key
            player.last_move_team_name = team_name
            player.last_move_source = source
        return

    if player.last_move_timestamp_unix is None or ts_unix >= player.last_move_timestamp_unix:
        player.last_move_timestamp_unix = ts_unix
        player.last_move_type = move_type
        player.last_move_team_key = team_key
        player.last_move_team_name = team_name
        player.last_move_source = source


def _build_player_summaries(
    league_dump: Dict[str, Any],
    draft_dump: Dict[str, Any],
    tx_dump: Dict[str, Any],
) -> List[PlayerSummary]:
    team_name_by_key = _build_team_name_map(league_dump)

    # Draft time (seconds since epoch) if available
    draft_time_unix: Optional[float] = None
    scoring = league_dump.get("scoring") or {}
    h2h = scoring.get("head_to_head") or {}
    if "draft_time" in h2h:
        try:
            draft_time_unix = float(h2h["draft_time"])
        except (TypeError, ValueError):
            draft_time_unix = None

    players: Dict[str, PlayerSummary] = {}

    def get_player(player_key: str) -> PlayerSummary:
        if player_key not in players:
            players[player_key] = PlayerSummary(player_key=player_key)
        return players[player_key]

    # 1) Seed from draft_dump
    for row in draft_dump.get("draft_results", []):
        if not isinstance(row, dict):
            continue
        player_key = row.get("player_key")
        if not player_key:
            continue

        team_key = row.get("team_key")
        team_name = row.get("team_name") or team_name_by_key.get(team_key)

        p = get_player(player_key)

        if p.drafted_team_key is None and team_key:
            p.drafted_team_key = team_key
            p.drafted_team_name = team_name

        # Treat the draft as the first "move" for this player.
        ts = draft_time_unix if draft_time_unix is not None else 0.0
        _update_last_move(
            player=p,
            ts_unix=ts,
            move_type="drafted",
            team_key=team_key,
            team_name=team_name,
            source="draft_dump",
        )

    # 2) Merge in transactions_dump
    for tx in tx_dump.get("transactions", []):
        if not isinstance(tx, dict):
            continue

        ts_raw = tx.get("timestamp_unix")
        try:
            ts_unix = float(ts_raw) if ts_raw is not None else None
        except (TypeError, ValueError):
            ts_unix = None

        moves = tx.get("moves") or []
        if not isinstance(moves, list):
            continue

        for mv in moves:
            if not isinstance(mv, dict):
                continue
            player_key = mv.get("player_key")
            if not player_key:
                continue

            p = get_player(player_key)

            # Capture a name if we don't have one yet.
            mv_name = mv.get("player_name")
            if isinstance(mv_name, str) and mv_name and not p.player_name:
                p.player_name = mv_name

            move_type_raw = mv.get("transaction_player_type") or tx.get("type")
            move_type = _normalize_move_type(move_type_raw)

            from_team_key = mv.get("from_team_key")
            to_team_key = mv.get("to_team_key")
            from_team_name = mv.get("from_team_name") or team_name_by_key.get(from_team_key)
            to_team_name = mv.get("to_team_name") or team_name_by_key.get(to_team_key)

            if move_type == "add":
                team_key = to_team_key or from_team_key
                team_name = to_team_name or from_team_name
            elif move_type == "drop":
                team_key = from_team_key or to_team_key
                team_name = from_team_name or to_team_name
            elif move_type == "trade":
                # Prefer the destination team as the "last team involved"
                team_key = to_team_key or from_team_key
                team_name = to_team_name or from_team_name
            else:
                # Fallback: whichever side we have.
                team_key = to_team_key or from_team_key
                team_name = to_team_name or from_team_name

            _update_last_move(
                player=p,
                ts_unix=ts_unix,
                move_type=move_type,
                team_key=team_key,
                team_name=team_name,
                source="transactions_dump",
            )

    # Return in a stable order
    out = list(players.values())
    out.sort(key=lambda p: (p.player_name or "", p.player_key))
    return out


# ---------------- Excel + manifest + latest.json ----------------


def _write_excel(players: List[PlayerSummary], path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(
            "ERROR: openpyxl is not installed but --to-excel was requested.\n"
            "Install it with 'pip install openpyxl' and try again.",
            file=sys.stderr,
        )
        sys.exit(1)

    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Players"

    headers = [
        "player_key",
        "player_name",
        "drafted_team_key",
        "drafted_team_name",
        "last_move_type",
        "last_move_team_key",
        "last_move_team_name",
        "last_move_source",
        "last_move_timestamp_unix",
    ]

    ws.append(headers)
    ws.freeze_panes = "A2"

    for p in players:
        ws.append(
            [
                p.player_key,
                p.player_name,
                p.drafted_team_key,
                p.drafted_team_name,
                p.last_move_type,
                p.last_move_team_key,
                p.last_move_team_name,
                p.last_move_source,
                p.last_move_timestamp_unix,
            ]
        )

    # Simple auto-width + autofilter
    last_col_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col_letter}1"

    for idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = 22

    wb.save(path)


def _write_manifest(
    paths: Paths,
    league_key: str,
    run_ts: RunTimestamps,
    cli_args: Dict[str, Any],
    processed_path: Path,
    excel_path: Optional[Path],
) -> Path:
    paths.manifest_dir.mkdir(parents=True, exist_ok=True)
    manifest_path = paths.manifest_dir / f"manifest.{run_ts.iso_stamp}.json"

    def rel(p: Path) -> str:
        return p.relative_to(paths.league_root).as_posix()

    files: Dict[str, Dict[str, Any]] = {}
    if processed_path.exists():
        files[rel(processed_path)] = {
            "size_bytes": processed_path.stat().st_size,
            "sha256": _sha256_of_file(processed_path),
        }
    if excel_path is not None and excel_path.exists():
        files[rel(excel_path)] = {
            "size_bytes": excel_path.stat().st_size,
            "sha256": _sha256_of_file(excel_path),
        }

    manifest = {
        "module": "rostered_players_list",
        "league_key": league_key,
        "_generated_unix": run_ts.unix,
        "_generated_iso_utc": run_ts.iso_utc,
        "_generated_iso_local": run_ts.iso_local,
        "cli_args": cli_args,
        "files": files,
    }

    _dump_json(manifest, manifest_path, pretty=True)
    return manifest_path


def _update_latest_meta(
    paths: Paths,
    league_key: str,
    processed_path: Path,
    excel_path: Optional[Path],
    run_ts: RunTimestamps,
) -> None:
    latest_path = paths.meta_dir / "latest.json"
    if latest_path.exists():
        try:
            with latest_path.open("r", encoding="utf-8") as f:
                latest = json.load(f)
        except Exception:
            latest = {}
    else:
        latest = {}

    latest["league_key"] = league_key

    rel_processed = processed_path.relative_to(paths.league_root).as_posix()
    rel_excel = (
        excel_path.relative_to(paths.league_root).as_posix()
        if excel_path is not None and excel_path.exists()
        else None
    )

    block: Dict[str, Any] = {"processed": rel_processed}
    if rel_excel is not None:
        block["excel"] = rel_excel

    latest["rostered_players_list"] = block
    latest["_updated_unix"] = run_ts.unix
    latest["_updated_iso_utc"] = run_ts.iso_utc

    _dump_json(latest, latest_path, pretty=True)


# ---------------- CLI + main ----------------


def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Build a consolidated player list from draft_dump and transactions_dump."
    )
    group = p.add_mutually_exclusive_group(required=True)
    group.add_argument("--league-key", help="Full league key, e.g. 465.l.22607")
    group.add_argument("--league-id", help="League ID, e.g. 22607 (paired with --game)")

    p.add_argument("--game", default="nhl", help="Game code (default: nhl)")
    p.add_argument("--pretty", action="store_true", help="Pretty-print JSON outputs")
    p.add_argument("--to-excel", action="store_true", help="Also write an Excel workbook")

    return p.parse_args()


def _resolve_league_key(args: argparse.Namespace) -> str:
    if args.league_key:
        return args.league_key
    return f"{args.game}.l.{args.league_id}"


def main() -> None:
    args = _parse_args()
    league_key = _resolve_league_key(args)

    paths = _paths_for_league(league_key)
    _ensure_dirs(paths)

    run_ts = make_run_timestamps()

    loaded = _load_inputs_for_league(paths, league_key)
    league_dump = loaded["league_dump"]
    draft_dump = loaded["draft_dump"]
    tx_dump = loaded["transactions_dump"]

    players = _build_player_summaries(league_dump, draft_dump, tx_dump)

    league_info = league_dump.get("league_info") or {}
    season = league_info.get("season")

    processed = {
        "league_key": league_info.get("league_key", league_key),
        "season": season,
        "generated_unix": run_ts.unix,
        "generated_iso_utc": run_ts.iso_utc,
        "generated_iso_local": run_ts.iso_local,
        "player_count": len(players),
        "players": [
            {
                "player_key": p.player_key,
                "player_name": p.player_name,
                "drafted_team_key": p.drafted_team_key,
                "drafted_team_name": p.drafted_team_name,
                "last_move_type": p.last_move_type,
                "last_move_team_key": p.last_move_team_key,
                "last_move_team_name": p.last_move_team_name,
                "last_move_source": p.last_move_source,
                "last_move_timestamp_unix": p.last_move_timestamp_unix,
            }
            for p in players
        ],
    }

    processed_path = paths.processed_dir / f"players.{run_ts.iso_stamp}.json"
    _dump_json(processed, processed_path, pretty=args.pretty)
    print(f"Wrote processed players JSON: {processed_path}")

    excel_path: Optional[Path] = None
    if args.to_excel:
        excel_path = paths.excel_dir / f"players.{run_ts.iso_stamp}.xlsx"
        _write_excel(players, excel_path)
        print(f"Wrote Excel workbook: {excel_path}")

    cli_args: Dict[str, Any] = {
        "league_key": league_key,
        "league_id": getattr(args, "league_id", None),
        "game": getattr(args, "game", None),
        "pretty": bool(getattr(args, "pretty", False)),
        "to_excel": bool(getattr(args, "to_excel", False)),
    }

    manifest_path = _write_manifest(
        paths=paths,
        league_key=league_key,
        run_ts=run_ts,
        cli_args=cli_args,
        processed_path=processed_path,
        excel_path=excel_path,
    )
    print(f"Wrote manifest: {manifest_path}")

    _update_latest_meta(
        paths=paths,
        league_key=league_key,
        processed_path=processed_path,
        excel_path=excel_path,
        run_ts=run_ts,
    )
    print(f"Updated latest metadata: {paths.meta_dir / 'latest.json'}")


if __name__ == "__main__":  # pragma: no cover
    main()
