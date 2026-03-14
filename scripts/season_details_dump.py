#!/usr/bin/env python3
from __future__ import annotations

"""season_details_dump: extract season-level details from a league dump and write a canonical season file.

This script reads a league's `_meta/latest.json` to find the latest `league_dump` processed JSON
and writes a normalized season details file to `exports/season_details/<season>/`.

Usage:
  python -m scripts.season_details_dump --league-key 465.l.22607 --season 2025 --pretty
"""

import argparse
import json
import sys
from pathlib import Path
from typing import Any, Dict

from src.util_time import make_run_timestamps
from src.config.env import get_export_dir


def _update_latest(season_root: Path, run_ts: Any, processed_rel: str, excel_rel: str = None) -> Path:
    """Update _meta/latest.json for the season."""
    meta_dir = season_root / "_meta"
    meta_dir.mkdir(parents=True, exist_ok=True)
    latest_path = meta_dir / "latest.json"

    if latest_path.exists():
        with latest_path.open("r", encoding="utf-8") as f:
            data = json.load(f)
    else:
        data = {"season": season_root.name}

    season_details = data.get("season_details", {})
    season_details["processed"] = processed_rel
    if excel_rel:
        season_details["excel"] = excel_rel

    data["season_details"] = season_details
    data["_updated_unix"] = run_ts.unix
    data["_updated_iso_utc"] = run_ts.iso_utc

    with latest_path.open("w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, sort_keys=True)

    return latest_path

def _to_excel(season: str, out_data: Dict[str, Any], xlsx_path: Path, run_ts: Any) -> None:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"Season {season}"
    ws.append(("Field", "Value"))
    
    for k, v in out_data.items():
        if v is None:
            v = ""
        ws.append((k, str(v)))
    
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:B{ws.max_row}"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)




def _load_latest_meta(league_root: Path) -> Dict[str, Any]:
    p = league_root / "_meta" / "latest.json"
    if not p.exists():
        print(f"ERROR: missing _meta/latest.json under {league_root}", file=sys.stderr)
        sys.exit(1)
    return json.loads(p.read_text(encoding="utf-8"))


def _read_processed_league(league_root: Path, latest: Dict[str, Any]) -> Dict[str, Any]:
    block = latest.get("league_dump")
    if not isinstance(block, dict) or "processed" not in block:
        print("ERROR: league_dump.processed not present in latest.json", file=sys.stderr)
        sys.exit(1)
    path = league_root / block["processed"]
    if not path.exists():
        print(f"ERROR: league_dump processed file not found: {path}", file=sys.stderr)
        sys.exit(1)
    return json.loads(path.read_text(encoding="utf-8"))


def main() -> None:
    ap = argparse.ArgumentParser(description="Write canonical season details from a league dump.")
    ap.add_argument("--season", required=True, help="Season year, e.g. 2025")
    ap.add_argument("--pretty", action="store_true")
    ap.add_argument("--to-excel", action="store_true", help="Also write an Excel workbook")
    args = ap.parse_args()
    season = str(args.season)

    # Build canonical, league-agnostic season details.
    run_ts = make_run_timestamps()
    out: Dict[str, Any] = {
        "source": "yahoo.game/nhl",
        "season": season,
        "_generated_unix": run_ts.unix,
        "_generated_iso_utc": run_ts.iso_utc,
        "_generated_iso_local": run_ts.iso_local,
        # placeholders for global metadata that can be added later
        "game_universe": None,
        "player_count_estimate": None,
    }

    season_root = get_export_dir() / season
    base = season_root / "season_details"
    base.mkdir(parents=True, exist_ok=True)
    out_path = base / f"season_details.{run_ts.iso_stamp}.json"
    with out_path.open("w", encoding="utf-8") as f:
        if args.pretty:
            json.dump(out, f, indent=2)
        else:
            json.dump(out, f, separators=(",", ":"))

    print(f"Wrote season details: {out_path}")
    
    processed_rel = out_path.relative_to(season_root).as_posix()
    excel_rel = None

    if args.to_excel:
        excel_path = base / f"season_details.{run_ts.iso_stamp}.xlsx"
        _to_excel(season, out, excel_path, run_ts)
        excel_rel = excel_path.relative_to(season_root).as_posix()
        print(f"Wrote Excel: {excel_path}")

    latest_path = _update_latest(season_root, run_ts, processed_rel, excel_rel)
    print(f"Updated latest.json: {latest_path}")


if __name__ == "__main__":
    main()
